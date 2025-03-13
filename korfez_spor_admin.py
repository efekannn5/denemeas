import sys
import sqlite3
from datetime import datetime, timedelta, date
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget, 
                            QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, 
                            QTableWidgetItem, QLineEdit, QLabel, QDialog, 
                            QFormLayout, QMessageBox, QComboBox, QTextEdit, QSpinBox, QFileDialog, QGroupBox,
                            QDateEdit, QGridLayout, QDialogButtonBox, QCompleter)
from PyQt5.QtCore import Qt, QTimer, QDate, QStringListModel
from PyQt5.QtGui import QDoubleValidator, QIcon, QColor  # QColor'ı buraya ekledik
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import pickle
import os.path
import io
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from fpdf import FPDF  # reportlab yerine
# QCompleter'ı QtWidgets'dan import et
from PyQt5.QtWidgets import QCompleter
import calendar

# Veritabanı bağlantısı ve tablo oluşturma
def create_database():
    conn = sqlite3.connect('korfez_spor.db')
    cursor = conn.cursor()
    
    # Gruplar tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS groups (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT
    )
    ''')
    
    # Öğrenciler tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        surname TEXT NOT NULL,
        tc_no TEXT,
        birth_date TEXT,
        blood_type TEXT,
        chronic_illness TEXT,
        phone TEXT,
        address TEXT,
        parent_name TEXT,
        parent_phone TEXT,
        emergency_contact TEXT,
        emergency_phone TEXT,
        fee REAL DEFAULT 0,
        group_id INTEGER,
        registration_date TEXT,
        payment_status TEXT DEFAULT 'Ödenmedi',
        payment_day INTEGER DEFAULT 1,
        FOREIGN KEY (group_id) REFERENCES groups (id)
    )
    ''')
    
    # Ödemeler tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        amount REAL,
        payment_date TEXT,
        payment_month INTEGER,
        payment_year INTEGER,
        status TEXT DEFAULT 'Ödenmedi',
        FOREIGN KEY (student_id) REFERENCES students (id)
    )
    ''')
    
    # Gelir tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS income (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        type TEXT,
        description TEXT,
        amount REAL
    )
    ''')
    
    # Notlar tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS notes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        note TEXT,
        date TEXT,
        FOREIGN KEY (student_id) REFERENCES students (id)
    )
    ''')
    
    # Eksik sütunları kontrol et ve ekle
    try:
        cursor.execute('ALTER TABLE students ADD COLUMN payment_day INTEGER DEFAULT 1')
    except sqlite3.OperationalError:
        pass
        
    try:
        cursor.execute('ALTER TABLE students ADD COLUMN payment_status TEXT DEFAULT "Ödenmedi"')
    except sqlite3.OperationalError:
        pass
        
    try:
        cursor.execute('ALTER TABLE students ADD COLUMN surname TEXT')
    except sqlite3.OperationalError:
        pass
        
    try:
        cursor.execute('ALTER TABLE students ADD COLUMN tc_no TEXT')
    except sqlite3.OperationalError:
        pass

    # Malzeme kategorileri tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS equipment_categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT
    )
    ''')
    
    # Malzemeler tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS equipment (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category_id INTEGER,
        name TEXT NOT NULL,
        size TEXT,
        purchase_price REAL,
        sale_price REAL,
        stock_quantity INTEGER DEFAULT 0,
        min_stock_level INTEGER DEFAULT 5,
        FOREIGN KEY (category_id) REFERENCES equipment_categories (id)
    )
    ''')
    
    # Sporcu malzeme kayıtları tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS student_equipment (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        equipment_id INTEGER,
        given_date TEXT,
        payment_status TEXT DEFAULT 'Ödenmedi',
        payment_amount REAL,
        FOREIGN KEY (student_id) REFERENCES students (id),
        FOREIGN KEY (equipment_id) REFERENCES equipment (id)
    )
    ''')
    
    # Stok hareketleri tablosu
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS stock_movements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        equipment_id INTEGER,
        movement_type TEXT,  -- 'IN' veya 'OUT'
        quantity INTEGER,
        date TEXT,
        description TEXT,
        FOREIGN KEY (equipment_id) REFERENCES equipment (id)
    )
    ''')

    # Varsayılan grupları ekle (eğer yoksa)
    cursor.execute("SELECT COUNT(*) FROM groups")
    if cursor.fetchone()[0] == 0:
        default_groups = [
            ("2007", "Default Açıklama"),
        ]
        cursor.executemany("INSERT INTO groups (name, description) VALUES (?, ?)", default_groups)
    
    # Varsayılan kategorileri ekle
    cursor.execute("SELECT COUNT(*) FROM equipment_categories")
    if cursor.fetchone()[0] == 0:
        default_categories = [
            ("Forma", "Takım formaları"),
            ("Şort", "Spor şortları"),
            ("Tozluk", "Futbol tozlukları"),
            ("Yağmurluk", "Yağmurluklar"),
            ("Diğer", "Diğer malzemeler")
        ]
        cursor.executemany(
            "INSERT INTO equipment_categories (name, description) VALUES (?, ?)", 
            default_categories
        )
    
    # İndeksler ekle
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_student_name ON students(name)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_payment_date ON payments(payment_date)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_payment_status ON students(payment_status)')
    
    # Malzeme tablosuna yeni sütunlar ekle
    try:
        cursor.execute('ALTER TABLE equipment ADD COLUMN purchase_price REAL DEFAULT 0')
    except sqlite3.OperationalError:
        pass
        
    try:
        cursor.execute('ALTER TABLE equipment ADD COLUMN sale_price REAL DEFAULT 0')
    except sqlite3.OperationalError:
        pass
        
    try:
        cursor.execute('ALTER TABLE equipment RENAME COLUMN price TO purchase_price')
    except sqlite3.OperationalError:
        pass

    # Mevcut price sütunundaki değerleri purchase_price ve sale_price'a kopyala
    try:
        cursor.execute('UPDATE equipment SET sale_price = purchase_price WHERE sale_price IS NULL')
    except sqlite3.OperationalError:
        pass
    
    conn.commit()
    conn.close()

def hesapla_aidat_tutari(tam_aidat, yeni_kayit=False):
    """
    Aidat tutarını hesaplar
    yeni_kayit: True ise yeni kayıt olan öğrenci için hesaplama yapar
    False ise mevcut öğrenci için hesaplama yapar
    """
    bugun = datetime.now()
    gun = bugun.day
    
    if yeni_kayit:
        # Yeni kayıt olan öğrenciler için kısmi ödeme
        if 1 <= gun <= 7:
            return tam_aidat  # Tam aidat
        elif 8 <= gun <= 15:
            return round(tam_aidat * 0.75)  # 3/4 aidat
        elif 16 <= gun <= 23:
            return round(tam_aidat * 0.5)  # 1/2 aidat
        else:
            return round(tam_aidat * 0.25)  # 1/4 aidat
    else:
        # Mevcut öğrenciler için ödeme kontrolü
        if 1 <= gun <= 7:
            return tam_aidat  # Normal ödeme dönemi
        else:
            # Gecikme zammı eklenebilir
            gecikme_zammi = round(tam_aidat * 0.1)  # %10 gecikme zammı
            return tam_aidat + gecikme_zammi

def kontrol_odeme_durumu(ogrenci_id):
    """Sporcunun bu ay için ödeme yapıp yapmadığını kontrol eder"""
    conn = sqlite3.connect('korfez_spor.db')
    cursor = conn.cursor()
    
    bugun = datetime.now()
    ay = bugun.month
    yil = bugun.year
    
    # Bu ay için ödeme kontrolü
    cursor.execute("""
        SELECT COUNT(*) FROM payments 
        WHERE student_id = ? AND payment_month = ? AND payment_year = ?
    """, (ogrenci_id, ay, yil))
    
    odeme_yapilmis = cursor.fetchone()[0] > 0
    
    # Öğrencinin kayıt tarihini kontrol et
    cursor.execute("SELECT registration_date FROM students WHERE id = ?", (ogrenci_id,))
    kayit_tarihi = datetime.strptime(cursor.fetchone()[0], '%Y-%m-%d')
    
    conn.close()
    
    # Eğer bu ay kayıt olduysa True döndür (ilk ay ödemesi yapılmış sayılır)
    if kayit_tarihi.month == bugun.month and kayit_tarihi.year == bugun.year:
        return True
        
    return odeme_yapilmis

def aidat_ode(self, ogrenci_id):
    try:
        bugun = datetime.now()
        
        # Eğer 1-7 arası değilse ve yeni kayıt değilse uyarı ver
        if bugun.day > 7:
            QMessageBox.warning(
                self,
                "Uyarı",
                "Aidat ödemeleri her ayın 1-7'si arasında yapılmalıdır!"
            )
            return
            
        # Normal ödeme işlemleri...
    except Exception as e:
        QMessageBox.critical(self, "Hata", f"Ödeme kaydedilirken hata oluştu: {str(e)}")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        create_database()
        self.init_ui()
        
        # Aylık yedekleme için timer
        self.backup_timer = QTimer()
        self.backup_timer.timeout.connect(self.check_monthly_backup)
        self.backup_timer.start(24 * 60 * 60 * 1000)  # Her gün kontrol et
        
        # İlk çalıştırmada kontrol et
        self.check_monthly_backup()
    
    def check_monthly_backup(self):
        """Her ayın başında ödemeleri kontrol et"""
        try:
            today = datetime.now()
            backup_folder = "backups"
            
            # Backup klasörü yoksa oluştur
            if not os.path.exists(backup_folder):
                os.makedirs(backup_folder)
            
            # Bu ay için yedek alınmış mı kontrol et
            backup_file = f"backup_{today.strftime('%Y_%m')}.db"
            backup_path = os.path.join(backup_folder, backup_file)
            
            if not os.path.exists(backup_path):
                # Veritabanını yedekle
                import shutil
                shutil.copy2('korfez_spor.db', backup_path)
                
                # Eski yedekleri kontrol et (6 aydan eski yedekleri sil)
                for old_backup in os.listdir(backup_folder):
                    backup_date = datetime.strptime(old_backup.split('_')[1].split('.')[0], '%Y_%m')
                    if (today - backup_date).days > 180:  # 6 ay
                        os.remove(os.path.join(backup_folder, old_backup))
                
                QMessageBox.information(
                    self,
                    "Yedekleme",
                    f"Veritabanı yedeklemesi oluşturuldu:\n{backup_file}"
                )
        
        except Exception as e:
            QMessageBox.warning(
                self,
                "Yedekleme Hatası",
                f"Yedekleme sırasında hata oluştu:\n{str(e)}"
            )

    def init_ui(self):
        # Logo ayarla
        try:
            self.setWindowIcon(QIcon('logo.png'))
        except:
            QMessageBox.warning(self, "Uyarı", "Logo dosyası (logo.png) bulunamadı!")
            
        # Pencere başlığını güncelle
        self.setWindowTitle("Körfezkent Spor Kulübü - Yönetim Paneli")
        self.setGeometry(100, 100, 1200, 800)
        
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        
        # Tab widget'ı sınıf değişkeni olarak tut
        self.tabs = QTabWidget()
        
        # Sekmeleri sınıf değişkenleri olarak tut
        self.athletes_tab = AthletesTab(self)  # StudentsTab -> AthletesTab
        self.payments_tab = PaymentsTab(self)
        self.notes_tab = NotesTab(self)
        self.unpaid_tab = UnpaidStudentsTab(self)
        self.accounting_tab = AccountingTab(self)
        self.equipment_tab = EquipmentTab(self)
        
        self.tabs.addTab(self.athletes_tab, "1-Sporcular")  # Öğrenciler -> Sporcular
        self.tabs.addTab(self.payments_tab, "2-Ödemeler")
        self.tabs.addTab(self.notes_tab, "4-Notlar")
        self.tabs.addTab(self.unpaid_tab, "Ödeme Bekleyenler")
        self.tabs.addTab(self.accounting_tab, "5-Muhasebe")
        self.tabs.addTab(self.equipment_tab, "6-Malzeme Yönetimi")
        
        layout.addWidget(self.tabs)
        
        # Powered by etiketi
        powered_by = QLabel()
        powered_by.setText('<a href="https://efekannefesoglu.com" style="color: #666; text-decoration: none;">Powered By Efekan NEFESOĞLU</a>')
        powered_by.setOpenExternalLinks(True)
        powered_by.setAlignment(Qt.AlignRight)
        
        font = powered_by.font()
        font.setPointSize(8)
        powered_by.setFont(font)
        
        powered_by.setStyleSheet("""
            QLabel {
                padding: 5px;
                margin: 5px;
            }
            QLabel:hover {
                color: #333;
            }
        """)
        
        layout.addWidget(powered_by)

    def refresh_all_tabs(self):
        """Tüm sekmeleri güncelle"""
        self.athletes_tab.load_students()
        self.payments_tab.load_payments()
        self.notes_tab.load_notes()
        self.unpaid_tab.load_unpaid_students()
        self.accounting_tab.load_data()
        self.equipment_tab.load_data()

class DatabaseConnection:
    def __init__(self):
        self.conn = None
        self.cursor = None

    def __enter__(self):
        self.conn = sqlite3.connect('korfez_spor.db')
        self.cursor = self.conn.cursor()
        return self.cursor

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.conn:
            if exc_type is None:
                self.conn.commit()
            else:
                self.conn.rollback()
            self.conn.close()

class AthletesTab(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        
        # Üst panel
        top_panel = QHBoxLayout()
        
        # Grup filtresi
        self.group_combo = QComboBox()
        self.group_combo.addItem("Tüm Gruplar", None)
        self.load_groups()
        top_panel.addWidget(QLabel("Yaş Grubu:"))
        top_panel.addWidget(self.group_combo)
        
        # Grup yönetim butonu
        manage_groups_btn = QPushButton("Grupları Yönet")
        top_panel.addWidget(manage_groups_btn)
        
        # Ara alanı
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Sporcu ara...")
        top_panel.addWidget(QLabel("Ara:"))
        top_panel.addWidget(self.search_input)
        
        # Butonlar
        add_btn = QPushButton("Yeni Sporcu")
        edit_btn = QPushButton("Düzenle")
        delete_btn = QPushButton("Sil")
        bulk_fee_btn = QPushButton("Toplu Aidat Güncelle")
        excel_btn = QPushButton("Excel'e Aktar")
        
        top_panel.addWidget(add_btn)
        top_panel.addWidget(edit_btn)
        top_panel.addWidget(delete_btn)
        top_panel.addWidget(bulk_fee_btn)
        top_panel.addWidget(excel_btn)
        top_panel.addStretch()
        
        # Tablo oluştur
        self.table = QTableWidget()
        self.table.setColumnCount(10)  # Sütun sayısını 10'a çıkar
        self.table.setHorizontalHeaderLabels([
            "Ad Soyad",
            "TC Kimlik No",
            "Doğum Tarihi", 
            "Kan Grubu",
            "Yaş Grubu",
            "Telefon",
            "Veli Adı",
            "Veli Telefon",
            "Acil Durum Kişisi",
            "Acil Durum Telefonu"
        ])
        
        # Tablo özellikleri
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        
        # Sütun genişlikleri
        self.table.setColumnWidth(0, 150)  # Ad Soyad
        self.table.setColumnWidth(1, 100)  # TC No
        self.table.setColumnWidth(2, 100)  # Doğum Tarihi
        self.table.setColumnWidth(3, 80)   # Kan Grubu
        self.table.setColumnWidth(4, 100)  # Yaş Grubu
        self.table.setColumnWidth(5, 100)  # Telefon
        self.table.setColumnWidth(6, 150)  # Veli Adı
        self.table.setColumnWidth(7, 100)  # Veli Telefon
        self.table.setColumnWidth(8, 150)  # Acil Durum Kişisi
        self.table.setColumnWidth(9, 100)  # Acil Durum Telefonu
        
        # Layout'a ekle
        layout.addLayout(top_panel)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        
        # Buton bağlantıları
        add_btn.clicked.connect(self.add_student)
        edit_btn.clicked.connect(self.edit_student)
        delete_btn.clicked.connect(self.delete_student)
        bulk_fee_btn.clicked.connect(self.show_bulk_fee_dialog)
        excel_btn.clicked.connect(self.export_to_excel)
        self.search_input.textChanged.connect(self.search_students)
        manage_groups_btn.clicked.connect(self.show_group_manager)
        self.group_combo.currentIndexChanged.connect(self.load_students)
        
        # İlk yükleme
        self.load_students()

    def load_groups(self):
        with DatabaseConnection() as cursor:
            cursor.execute("SELECT id, name FROM groups ORDER BY name")
            groups = cursor.fetchall()
            for group in groups:
                self.group_combo.addItem(group[1], group[0])

    def load_students(self):
        try:
            self.table.setRowCount(0)
            group_id = self.group_combo.currentData()
            search_text = self.search_input.text().strip()
            
            with DatabaseConnection() as cursor:
                if group_id:
                    cursor.execute("""
                        SELECT 
                            s.name || ' ' || s.surname as full_name,
                            s.tc_no,
                            s.birth_date,
                            s.blood_type,
                            g.name as group_name,
                            s.phone,
                            s.parent_name,
                            s.parent_phone,
                            s.emergency_contact,
                            s.emergency_phone
                        FROM students s
                        LEFT JOIN groups g ON s.group_id = g.id
                        WHERE s.group_id = ?
                        ORDER BY s.name
                    """, (group_id,))
                else:
                    cursor.execute("""
                        SELECT 
                            s.name || ' ' || s.surname as full_name,
                            s.tc_no,
                            s.birth_date,
                            s.blood_type,
                            g.name as group_name,
                            s.phone,
                            s.parent_name,
                            s.parent_phone,
                            s.emergency_contact,
                            s.emergency_phone
                        FROM students s
                        LEFT JOIN groups g ON s.group_id = g.id
                        ORDER BY s.name
                    """)
                
                students = cursor.fetchall()
                
                for student in students:
                    if search_text and search_text.lower() not in student[0].lower():  # full_name kontrolü
                        continue
                        
                    row = self.table.rowCount()
                    self.table.insertRow(row)
                    
                    for col, value in enumerate(student):
                        # Doğum tarihini formatla
                        if col == 2 and value:  # birth_date sütunu
                            try:
                                date = datetime.strptime(value, '%Y-%m-%d')
                                value = date.strftime('%d/%m/%Y')
                            except:
                                pass
                        
                        self.table.setItem(row, col, QTableWidgetItem(str(value or '')))
                    
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Sporcular yüklenirken hata oluştu: {str(e)}")

    def hesapla_aidat(self):
        try:
            tam_aidat = int(self.aidat_entry.text())
            bugun = datetime.now()
            gun = bugun.day
            
            # Haftayı hesapla (1-7: 1.hafta, 8-15: 2.hafta, 16-23: 3.hafta, 24-31: 4.hafta)
            if 1 <= gun <= 7:
                hafta = 1
                oran = 1.0  # tam aidat
            elif 8 <= gun <= 15:
                hafta = 2
                oran = 0.75  # 3/4 aidat
            elif 16 <= gun <= 23:
                hafta = 3
                oran = 0.5  # 1/2 aidat
            else:
                hafta = 4
                oran = 0.25  # 1/4 aidat
            
            odenecek_aidat = round(tam_aidat * oran)
            
            self.hesaplanan_aidat_label.setText(
                f"{odenecek_aidat} TL\n"
                f"({hafta}. hafta - {oran*100}% ödeme)"
            )
            
        except ValueError:
            QMessageBox.warning(self, "Hata", "Geçerli bir aidat tutarı girin!")

    def yeni_kayit(self):
        try:
            # Temel bilgileri al
            ad_soyad = self.ad_soyad_entry.text().strip()
            tc_no = self.tc_no_entry.text().strip()
            dogum_tarihi = self.dogum_tarihi_entry.text().strip()
            telefon = self.telefon_entry.text().strip()
            adres = self.adres_entry.toPlainText().strip()
            veli_ad = self.veli_ad_entry.text().strip()
            veli_tel = self.veli_tel_entry.text().strip()
            grup_id = self.grup_combo.currentData()
            
            tam_aidat = int(self.aidat_entry.text())
            bugun = datetime.now()
            gun = bugun.day
            
            # Haftayı ve oranı hesapla
            if 1 <= gun <= 7:
                oran = 1.0
                hafta = 1
            elif 8 <= gun <= 15:
                oran = 0.75
                hafta = 2
            elif 16 <= gun <= 23:
                oran = 0.5
                hafta = 3
            else:
                oran = 0.25
                hafta = 4
            
            odenecek_aidat = round(tam_aidat * oran)
            
            # Onay mesajı
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle("Aidat Bilgisi")
            msg.setText(
                f"Kayıt tarihi: {bugun.strftime('%d/%m/%Y')}\n"
                f"{hafta}. hafta\n"
                f"Normal aidat: {tam_aidat} TL\n"
                f"Ödenecek aidat: {odenecek_aidat} TL ({oran*100}% ödeme)\n\n"
                f"Sporcuyu kaydetmek ve aidatı tahsil etmek istiyor musunuz?"
            )
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            
            if msg.exec_() == QMessageBox.No:
                return
                
            conn = sqlite3.connect('korfez_spor.db')
            cursor = conn.cursor()
            
            # Öğrenciyi kaydet
            cursor.execute("""
                INSERT INTO students 
                (name, tc_no, birth_date, phone, address, registration_date, fee, group_id, parent_name, parent_phone) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                ad_soyad, tc_no, dogum_tarihi, telefon, adres, 
                bugun.strftime('%Y-%m-%d'),
                tam_aidat,  # Normal aidat tutarını kaydet
                grup_id, veli_ad, veli_tel
            ))
            
            # Yeni eklenen öğrencinin ID'sini al
            ogrenci_id = cursor.lastrowid
            
            # İlk aidat ödemesini kaydet
            cursor.execute("""
                INSERT INTO payments 
                (student_id, amount, payment_date, payment_month, payment_year, status) 
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                ogrenci_id,
                odenecek_aidat,
                bugun.strftime('%Y-%m-%d'),
                bugun.month,
                bugun.year,
                'Ödendi'
            ))
            
            conn.commit()
            conn.close()
            
            # Başarılı mesajı göster
            QMessageBox.information(
                self,
                "Başarılı",
                f"Sporcu kaydedildi ve ilk aidat ödemesi ({odenecek_aidat} TL) alındı."
            )
            
            # Formu temizle
            self.clear_form()
            
            # Tabloları güncelle
            self.load_students()
            self.parent().payments_tab.load_payments()
            self.parent().accounting_tab.load_data()
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kayıt sırasında hata oluştu: {str(e)}")

    def add_student(self):
        dialog = StudentDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            # Kayıt işlemi artık StudentDialog'da yapılıyor, 
            # burada sadece listeyi yenilememiz yeterli
            self.load_students()

    def edit_student(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir sporcu seçin")
            return
        
        student_id = int(self.table.item(selected_items[0].row(), 0).text())
        dialog = StudentDialog(self, student_id)
        
        if dialog.exec_() == QDialog.Accepted:
            # Güncelleme işlemi artık StudentDialog'da yapılıyor
            self.load_students()

    def delete_student(self):
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir öğrenci seçin")
            return
            
        reply = QMessageBox.question(self, "Onay", 
                                   "Bu sporcuyu silmek istediğinizden emin misiniz?",
                                   QMessageBox.Yes | QMessageBox.No)
                                   
        if reply == QMessageBox.Yes:
            student_id = self.table.item(current_row, 0).text()
            with DatabaseConnection() as cursor:
                cursor.execute('''
                    DELETE FROM students
                    WHERE id = ?
                ''', (student_id,))
            self.parent.refresh_all_tabs()

    def search_students(self, text):
        # Arama fonksiyonunu güncelle
        text = text.lower()
        for row in range(self.table.rowCount()):
            show = False
            # Tüm sütunlarda ara
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and text in item.text().lower():
                    show = True
                    break
            self.table.setRowHidden(row, not show)

    def show_bulk_fee_dialog(self):
        dialog = BulkFeeUpdateDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.parent.refresh_all_tabs()

    def export_to_excel(self):
        import_excel_modules()
        try:
            # Seçili grubu al
            selected_group_id = self.group_combo.currentData()
            selected_group_name = self.group_combo.currentText()
            
            # Dosya adı için tarih
            current_date = datetime.now().strftime('%Y%m%d')
            
            # Varsayılan dosya adı
            default_filename = f"ogrenciler_{current_date}.xlsx"
            if selected_group_id:
                default_filename = f"{selected_group_name}_ogrenciler_{current_date}.xlsx"
            
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Dosyasını Kaydet",
                default_filename,
                "Excel Dosyası (*.xlsx)"
            )
            
            if file_name:
                with DatabaseConnection() as cursor:
                    # SQL sorgusunu güncelle - ödeme bilgilerini çıkar
                    query = """
                        SELECT 
                            s.name || ' ' || s.surname as ad_soyad,
                            s.tc_no,
                            s.birth_date as dogum_tarihi,
                            s.blood_type as kan_grubu,
                            g.name as grup,
                            s.phone as telefon,
                            s.parent_name as veli_adi,
                            s.parent_phone as veli_telefon,
                            s.emergency_contact as acil_durum_yakini,
                            s.emergency_phone as acil_durum_telefon,
                            s.address as adres,
                            s.chronic_illness as kronik_hastalik,
                            s.registration_date as kayit_tarihi
                        FROM students s
                        LEFT JOIN groups g ON s.group_id = g.id
                    """
                    
                    if selected_group_id:
                        query += " WHERE s.group_id = ?"
                        cursor.execute(query, (selected_group_id,))
                    else:
                        cursor.execute(query)
                    
                    students = cursor.fetchall()
                    
                    # Sütun başlıklarını güncelle - ödeme ile ilgili sütunları çıkar
                    columns = [
                        'Ad Soyad', 
                        'TC No', 
                        'Doğum Tarihi', 
                        'Kan Grubu', 
                        'Grup', 
                        'Telefon', 
                        'Veli Adı', 
                        'Veli Telefon',
                        'Acil Durum Yakını', 
                        'Acil Durum Telefon', 
                        'Adres', 
                        'Kronik Hastalık',
                        'Kayıt Tarihi'
                    ]
                    
                    df = pd.DataFrame(students, columns=columns)
                    
                    # Tarih formatlarını düzenle
                    df['Kayıt Tarihi'] = pd.to_datetime(df['Kayıt Tarihi']).dt.strftime('%d/%m/%Y')
                    df['Doğum Tarihi'] = pd.to_datetime(df['Doğum Tarihi']).dt.strftime('%d/%m/%Y')
                    
                    # Excel'e kaydet
                    df.to_excel(file_name, index=False, sheet_name='Öğrenciler')
                    
                    # Excel formatlaması
                    wb = openpyxl.load_workbook(file_name)
                    ws = wb.active
                    
                    # Başlık stilini ayarla
                    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    
                    # Tüm sütunları otomatik genişlik yap
                    for column in ws.columns:
                        max_length = 0
                        column_letter = openpyxl.utils.get_column_letter(column[0].column)
                        
                        for cell in column:
                            if cell.row == 1:  # Başlık satırı
                                cell.font = Font(bold=True, color='FFFFFF')
                                cell.fill = header_fill
                                cell.alignment = Alignment(horizontal='center')
                            else:
                                cell.alignment = Alignment(horizontal='center')
                            
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    wb.save(file_name)
                    
                    QMessageBox.information(
                        self,
                        "Başarılı",
                        f"Öğrenci listesi Excel dosyasına aktarıldı:\n{file_name}"
                    )
                    
        except Exception as e:
            QMessageBox.critical(
                self,
                "Hata",
                f"Excel dosyası oluşturulurken bir hata oluştu:\n{str(e)}"
            )

    def show_group_manager(self):
        dialog = AgeGroupManagerDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_groups()  # Grupları yeniden yükle
            self.load_students()  # Öğrencileri yeniden yükle

class StudentDialog(QDialog):
    def __init__(self, parent=None, student_id=None):
        super().__init__(parent)
        self.student_id = student_id
        self.setWindowTitle("Sporcu Ekle/Düzenle")
        self.setModal(True)
        self.setMinimumWidth(500)
        
        layout = QVBoxLayout()
        
        # Kişisel Bilgiler Grubu
        personal_group = QGroupBox("Kişisel Bilgiler")
        personal_layout = QFormLayout()
        
        self.name_input = QLineEdit()
        self.surname_input = QLineEdit()
        self.tc_no_input = QLineEdit()  # TC No alanı eklendi
        self.birth_date_input = QDateEdit()
        self.birth_date_input.setCalendarPopup(True)  # Takvim popup'ı ekle
        self.birth_date_input.setDisplayFormat("dd/MM/yyyy")  # Görüntüleme formatı
        
        # Minimum ve maximum tarihleri ayarla (örnek: 5-18 yaş arası)
        current_date = QDate.currentDate()
        self.birth_date_input.setMaximumDate(current_date.addYears(-5))  # En küçük 5 yaş
        self.birth_date_input.setMinimumDate(current_date.addYears(-18))  # En büyük 18 yaş
        
        # Varsayılan tarihi ayarla
        self.birth_date_input.setDate(current_date.addYears(-10))  # Varsayılan 10 yaş
        
        self.blood_type_combo = QComboBox()
        self.blood_type_combo.addItems(["", "A Rh+", "A Rh-", "B Rh+", "B Rh-", "AB Rh+", "AB Rh-", "0 Rh+", "0 Rh-"])
        self.chronic_illness_input = QTextEdit()
        self.chronic_illness_input.setPlaceholderText("Varsa kalıcı hastalıkları yazın...")
        self.chronic_illness_input.setMaximumHeight(60)
        
        personal_layout.addRow("Ad:", self.name_input)
        personal_layout.addRow("Soyad:", self.surname_input)
        personal_layout.addRow("TC No:", self.tc_no_input)  # TC No form alanı eklendi
        personal_layout.addRow("Doğum Tarihi:", self.birth_date_input)
        personal_layout.addRow("Kan Grubu:", self.blood_type_combo)
        personal_layout.addRow("Kalıcı Hastalıklar:", self.chronic_illness_input)
        
        # Yaş grubu seçimi
        self.group_combo = QComboBox()
        self.group_combo.addItem("Grup Seçiniz", None)
        self.load_groups()
        personal_layout.addRow("Yaş Grubu:", self.group_combo)
        
        personal_group.setLayout(personal_layout)
        layout.addWidget(personal_group)
        
        # İletişim Bilgileri Grubu
        contact_group = QGroupBox("İletişim Bilgileri")
        contact_layout = QFormLayout()
        
        self.phone_input = QLineEdit()
        self.address_input = QTextEdit()
        self.address_input.setMaximumHeight(60)
        self.parent_name_input = QLineEdit()
        self.parent_phone_input = QLineEdit()
        self.emergency_contact_input = QLineEdit()
        self.emergency_phone_input = QLineEdit()
        
        contact_layout.addRow("Telefon:", self.phone_input)
        contact_layout.addRow("Adres:", self.address_input)
        contact_layout.addRow("Veli Adı:", self.parent_name_input)
        contact_layout.addRow("Veli Telefonu:", self.parent_phone_input)
        contact_layout.addRow("Acil Durum Kişisi:", self.emergency_contact_input)
        contact_layout.addRow("Acil Durum Telefonu:", self.emergency_phone_input)
        contact_group.setLayout(contact_layout)
        layout.addWidget(contact_group)
        
        # Aidat Bilgileri Grubu
        payment_group = QGroupBox("Aidat Bilgileri")
        payment_layout = QFormLayout()
        
        self.aidat_entry = QLineEdit()
        self.aidat_entry.setPlaceholderText("Aylık aidat tutarı")
        self.aidat_entry.setValidator(QDoubleValidator(0, 100000, 2))
        
        hesapla_btn = QPushButton("Aidat Hesapla")
        hesapla_btn.clicked.connect(self.hesapla_aidat)
        
        self.hesaplanan_aidat_label = QLabel()
        self.hesaplanan_tutar = 0  # Hesaplanan tutarı saklamak için
        
        payment_layout.addRow("Aidat Tutarı (TL):", self.aidat_entry)
        payment_layout.addRow(hesapla_btn)
        payment_layout.addRow("Ödenecek Tutar:", self.hesaplanan_aidat_label)
        
        payment_group.setLayout(payment_layout)
        layout.addWidget(payment_group)
        
        # Butonlar
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        self.setLayout(layout)
        
        if student_id:
            self.load_student_data()

    def load_groups(self):
        try:
            with DatabaseConnection() as cursor:
                cursor.execute('SELECT id, name FROM groups ORDER BY name')
                groups = cursor.fetchall()
                for group in groups:
                    self.group_combo.addItem(group[1], group[0])
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Gruplar yüklenirken hata oluştu: {str(e)}")

    def load_student_data(self):
        with DatabaseConnection() as cursor:
            cursor.execute('''
                SELECT 
                    name, surname, tc_no, phone, blood_type,
                    parent_name, parent_phone, address,
                    chronic_illness, emergency_contact,
                    emergency_phone, fee, group_id, birth_date
                FROM students 
                WHERE id = ?
            ''', (self.student_id,))
            student = cursor.fetchone()
            
            if student:
                self.name_input.setText(student[0])
                self.surname_input.setText(student[1])
                self.tc_no_input.setText(student[2] or '')  # TC No yükleme eklendi
                self.phone_input.setText(student[3] or '')
                self.blood_type_combo.setCurrentText(student[4] or '')
                self.parent_name_input.setText(student[5] or '')
                self.parent_phone_input.setText(student[6] or '')
                self.address_input.setText(student[7] or '')
                self.chronic_illness_input.setText(student[8] or '')
                self.emergency_contact_input.setText(student[9] or '')
                self.emergency_phone_input.setText(student[10] or '')
                self.aidat_entry.setText(str(student[11] or ''))
                self.group_combo.setCurrentIndex(
                    self.group_combo.findData(student[12])  # group_id
                )
                
                # Doğum tarihini ayarla
                if student[13]:  # birth_date
                    birth_date = QDate.fromString(student[13], "yyyy-MM-dd")
                    self.birth_date_input.setDate(birth_date)
            
    def hesapla_aidat(self):
        try:
            tam_aidat = float(self.aidat_entry.text())
            bugun = datetime.now()
            gun = bugun.day
            
            # Haftayı hesapla
            if 1 <= gun <= 7:
                hafta = 1
                oran = 1.0
            elif 8 <= gun <= 15:
                hafta = 2
                oran = 0.75
            elif 16 <= gun <= 23:
                hafta = 3
                oran = 0.5
            else:
                hafta = 4
                oran = 0.25
            
            self.hesaplanan_tutar = round(tam_aidat * oran)
            
            self.hesaplanan_aidat_label.setText(
                f"{self.hesaplanan_tutar} TL\n"
                f"({hafta}. hafta - %{int(oran*100)} ödeme)"
            )
            
        except ValueError:
            QMessageBox.warning(self, "Hata", "Geçerli bir aidat tutarı girin!")

    def accept(self):
        try:
            # Temel kontroller
            name = self.name_input.text().strip()
            surname = self.surname_input.text().strip()
            
            if not name or not surname:
                QMessageBox.warning(self, "Hata", "Ad ve soyad alanları zorunludur!")
                return
            
            if not self.aidat_entry.text() or not self.hesaplanan_tutar:
                QMessageBox.warning(self, "Hata", "Lütfen aidat tutarını hesaplayın!")
                return
            
            # Veritabanına kaydet
            with DatabaseConnection() as cursor:
                if self.student_id:  # Güncelleme
                    cursor.execute('''
                        UPDATE students SET 
                            name=?, surname=?, tc_no=?, birth_date=?, blood_type=?,
                            chronic_illness=?, phone=?, address=?,
                            parent_name=?, parent_phone=?,
                            emergency_contact=?, emergency_phone=?,
                            fee=?, group_id=?
                        WHERE id=?
                    ''', (
                        name, surname,  # Ad ve soyad ayrı ayrı
                        self.tc_no_input.text(),
                        self.birth_date_input.date().toString("yyyy-MM-dd"),  # Tarihi veritabanı formatına çevir
                        self.blood_type_combo.currentText(),
                        self.chronic_illness_input.toPlainText(),
                        self.phone_input.text(),
                        self.address_input.toPlainText(),
                        self.parent_name_input.text(),
                        self.parent_phone_input.text(),
                        self.emergency_contact_input.text(),
                        self.emergency_phone_input.text(),
                        float(self.aidat_entry.text()),  # Tam aidat tutarı
                        self.group_combo.currentData(),
                        self.student_id
                    ))
                else:  # Yeni kayıt
                    cursor.execute('''
                        INSERT INTO students (
                            name, surname, tc_no, birth_date, blood_type,
                            chronic_illness, phone, address,
                            parent_name, parent_phone,
                            emergency_contact, emergency_phone,
                            fee, group_id, registration_date,
                            payment_status  # Ödeme durumunu ekle
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        name, surname,
                        self.tc_no_input.text(),
                        self.birth_date_input.date().toString("yyyy-MM-dd"),
                        self.blood_type_combo.currentText(),
                        self.chronic_illness_input.toPlainText(),
                        self.phone_input.text(),
                        self.address_input.toPlainText(),
                        self.parent_name_input.text(),
                        self.parent_phone_input.text(),
                        self.emergency_contact_input.text(),
                        self.emergency_phone_input.text(),
                        float(self.aidat_entry.text()),
                        self.group_combo.currentData(),
                        datetime.now().strftime('%Y-%m-%d'),
                        'Ödenmedi'  # Başlangıçta ödenmedi olarak işaretle
                    ))
                    
                    student_id = cursor.lastrowid
                    
                    # İlk aidat ödemesini kaydetme kısmını kaldır
                    # Bunun yerine PaymentDialog ile ödeme alınacak
            
            # Tüm tabloları yenile
            if isinstance(self.parent(), AthletesTab):
                main_window = self.parent().parent
                main_window.refresh_all_tabs()
            
            super().accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kayıt sırasında hata oluştu: {str(e)}")

class AdminsTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Üst panel
        top_panel = QHBoxLayout()
        
        # Butonlar
        add_btn = QPushButton("Yeni Admin")
        edit_btn = QPushButton("Düzenle")
        delete_btn = QPushButton("Sil")
        
        top_panel.addWidget(add_btn)
        top_panel.addWidget(edit_btn)
        top_panel.addWidget(delete_btn)
        top_panel.addStretch()
        
        # Tablo
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(
            ["ID", "E-posta", "Şifre"]
        )
        
        layout.addLayout(top_panel)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        
        # Buton bağlantıları
        add_btn.clicked.connect(self.add_admin)
        edit_btn.clicked.connect(self.edit_admin)
        delete_btn.clicked.connect(self.delete_admin)
        
        # Adminleri yükle
        self.load_admins()
        
    def add_admin(self):
        dialog = AdminDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            # Firebase'e kaydet
            admin_ref = self.db.child('admins').push()
            admin_ref.set({
                'email': dialog.email_input.text(),
                'password': dialog.password_input.text()
            })
            self.load_admins()

    def edit_admin(self):
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir admin seçin")
            return
            
        admin_id = self.table.item(current_row, 0).text()
        dialog = AdminDialog(self)
        
        dialog.email_input.setText(self.table.item(current_row, 1).text())
        dialog.password_input.setText(self.table.item(current_row, 2).text())
        
        if dialog.exec_() == QDialog.Accepted:
            self.db.child('admins').child(admin_id).update({
                'email': dialog.email_input.text(),
                'password': dialog.password_input.text()
            })
            self.load_admins()

    def delete_admin(self):
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir admin seçin")
            return
            
        reply = QMessageBox.question(self, "Onay", 
                                   "Bu admini silmek istediğinizden emin misiniz?",
                                   QMessageBox.Yes | QMessageBox.No)
                                   
        if reply == QMessageBox.Yes:
            admin_id = self.table.item(current_row, 0).text()
            self.db.child('admins').child(admin_id).delete()
            self.load_admins()

    def load_admins(self):
        self.table.setRowCount(0)
        admins = self.db.child('admins').get()
        
        if admins:
            for admin_id, data in admins.items():
                row = self.table.rowCount()
                self.table.insertRow(row)
                self.table.setItem(row, 0, QTableWidgetItem(admin_id))
                self.table.setItem(row, 1, QTableWidgetItem(data['email']))
                self.table.setItem(row, 2, QTableWidgetItem(data['password']))

class AdminDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Admin Bilgileri")
        self.setModal(True)
        
        layout = QFormLayout()
        
        self.email_input = QLineEdit()
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        
        layout.addRow("E-posta:", self.email_input)
        layout.addRow("Şifre:", self.password_input)
        
        buttons = QHBoxLayout()
        save_btn = QPushButton("Kaydet")
        cancel_btn = QPushButton("İptal")
        
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        
        buttons.addWidget(save_btn)
        buttons.addWidget(cancel_btn)
        
        layout.addRow(buttons)
        self.setLayout(layout)

class PaymentsTab(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.init_ui()
        
        # Her gün kontrol et
        self.timer = QTimer()
        self.timer.timeout.connect(self.check_payments)
        self.timer.start(24 * 60 * 60 * 1000)  # 24 saat
        
        # Başlangıçta kontrol et
        self.check_payments()
    
    def check_payments(self):
        """Her ayın başında ödemeleri kontrol et"""
        bugun = datetime.now()
        
        with DatabaseConnection() as cursor:
            # Önce tüm öğrencilerin bu ayki ödeme durumunu kontrol et
            cursor.execute("""
                UPDATE students 
                SET payment_status = CASE
                    WHEN id IN (
                        SELECT student_id FROM payments 
                        WHERE payment_month = ? AND payment_year = ?
                    ) THEN 'Ödendi'
                    WHEN ? > 7 THEN 'GECİKMİŞ ÖDEME'
                    ELSE 'Ödenmedi'
                END
            """, (bugun.month, bugun.year, bugun.day))
            
            # Yeni kayıt olanları kontrol et
            cursor.execute("""
                UPDATE students 
                SET payment_status = 'Ödendi'
                WHERE strftime('%Y-%m', registration_date) = strftime('%Y-%m', 'now')
                AND id IN (
                    SELECT student_id FROM payments 
                    WHERE payment_month = ? AND payment_year = ?
                )
            """, (bugun.month, bugun.year))
        
        self.load_payments()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Üst panel
        top_panel = QHBoxLayout()
        
        # Ay ve yıl seçici
        self.month_combo = QComboBox()
        self.month_combo.addItems(["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                                 "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"])
        current_month = datetime.now().month
        self.month_combo.setCurrentIndex(current_month - 1)
        
        self.year_spin = QSpinBox()
        self.year_spin.setRange(2020, 2100)
        self.year_spin.setValue(datetime.now().year)
        
        top_panel.addWidget(QLabel("Ay:"))
        top_panel.addWidget(self.month_combo)
        top_panel.addWidget(QLabel("Yıl:"))
        top_panel.addWidget(self.year_spin)
        
        # Butonlar
        add_btn = QPushButton("Ödeme Ekle")
        report_btn = QPushButton("Aylık Rapor")
        
        top_panel.addWidget(add_btn)
        top_panel.addWidget(report_btn)
        
        # Excel export butonları için yeni panel
        excel_panel = QHBoxLayout()
        excel_panel.addWidget(QLabel("Excel Raporu:"))
        
        last_month_btn = QPushButton("Son 1 Ay")
        last_3months_btn = QPushButton("Son 3 Ay")
        all_time_btn = QPushButton("Tüm Zamanlar")
        
        last_month_btn.clicked.connect(lambda: self.export_to_excel("last_month"))
        last_3months_btn.clicked.connect(lambda: self.export_to_excel("last_3months"))
        all_time_btn.clicked.connect(lambda: self.export_to_excel("all_time"))
        
        excel_panel.addWidget(last_month_btn)
        excel_panel.addWidget(last_3months_btn)
        excel_panel.addWidget(all_time_btn)
        excel_panel.addStretch()
        
        # Özet panel
        summary_panel = QHBoxLayout()
        self.total_label = QLabel("Toplam Gelir: 0 TL")
        self.paid_count_label = QLabel("Ödeme Yapan: 0")
        self.unpaid_count_label = QLabel("Ödeme Yapmayan: 0")
        
        summary_panel.addWidget(self.total_label)
        summary_panel.addWidget(self.paid_count_label)
        summary_panel.addWidget(self.unpaid_count_label)
        
        # Tablo
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(
            ["Öğrenci", "Aidat Günü", "Son Ödeme", "Durum", "Kalan Gün", "Tutar"]
        )
        
        layout.addLayout(top_panel)
        layout.addLayout(excel_panel)
        layout.addLayout(summary_panel)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        
        # Bağlantılar
        add_btn.clicked.connect(self.add_payment)
        report_btn.clicked.connect(self.generate_monthly_report)
        self.month_combo.currentIndexChanged.connect(self.load_payments)
        self.year_spin.valueChanged.connect(self.load_payments)
        
        self.load_payments()

    def load_payments(self):
        try:
            self.table.setRowCount(0)
            selected_month = self.month_combo.currentIndex() + 1
            selected_year = self.year_spin.value()
            
            with DatabaseConnection() as cursor:
                # Ödemeleri yükle
                cursor.execute("""
                    SELECT 
                        s.name || ' ' || s.surname as student_name,
                        s.payment_day,
                        MAX(p.payment_date) as last_payment,
                        s.payment_status,
                        s.fee
                    FROM students s
                    LEFT JOIN payments p ON s.id = p.student_id 
                        AND p.payment_month = ? 
                        AND p.payment_year = ?
                    GROUP BY s.id
                    ORDER BY student_name
                """, (selected_month, selected_year))
                
                payments = cursor.fetchall()
                
                # Özet bilgileri hesapla
                total_amount = 0
                paid_count = 0
                unpaid_count = 0
                
                for payment in payments:
                    row = self.table.rowCount()
                    self.table.insertRow(row)
                    
                    student_name = payment[0]
                    payment_day = payment[1] or 1
                    last_payment = payment[2] or "Ödeme yapılmamış"
                    status = payment[3] or "Ödenmedi"
                    fee = payment[4] or 0
                    
                    # Kalan gün hesapla
                    if status == "Ödenmedi":
                        today = datetime.now()
                        if today.day <= payment_day:
                            remaining_days = payment_day - today.day
                        else:
                            remaining_days = 0
                    else:
                        remaining_days = "-"
                    
                    # İstatistikleri güncelle
                    if status == "Ödendi":
                        paid_count += 1
                        total_amount += fee
                    else:
                        unpaid_count += 1
                    
                    # Tabloyu doldur
                    self.table.setItem(row, 0, QTableWidgetItem(student_name))
                    self.table.setItem(row, 1, QTableWidgetItem(str(payment_day)))
                    self.table.setItem(row, 2, QTableWidgetItem(str(last_payment)))
                    
                    status_item = QTableWidgetItem(status)
                    status_item.setForeground(Qt.green if status == "Ödendi" else Qt.red)
                    self.table.setItem(row, 3, status_item)
                    
                    self.table.setItem(row, 4, QTableWidgetItem(str(remaining_days)))
                    self.table.setItem(row, 5, QTableWidgetItem(f"{fee} TL"))
                
                # Özet bilgileri güncelle
                self.total_label.setText(f"Toplam Gelir: {total_amount} TL")
                self.paid_count_label.setText(f"Ödeme Yapan: {paid_count}")
                self.unpaid_count_label.setText(f"Ödeme Yapmayan: {unpaid_count}")
                
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ödemeler yüklenirken hata oluştu: {str(e)}")

    def add_payment(self):
        dialog = PaymentDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            try:
                student_id = dialog.student_combo.currentData()
                amount = float(dialog.amount_input.text())
                payment_date = dialog.date_input.date()
                
                with DatabaseConnection() as cursor:
                    # Ödemeyi kaydet
                    cursor.execute('''
                        INSERT INTO payments (
                            student_id, amount, payment_date, 
                            payment_month, payment_year, status
                        )
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (
                        student_id,
                        amount,
                        payment_date.toString("yyyy-MM-dd"),
                        payment_date.month(),
                        payment_date.year(),
                        'Ödendi'
                    ))
                    
                    # Öğrencinin ödeme durumunu güncelle
                    cursor.execute('''
                        UPDATE students 
                        SET payment_status = 'Ödendi'
                        WHERE id = ?
                    ''', (student_id,))
                
                self.load_payments()  # Tabloyu güncelle
                
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Hata",
                    f"Ödeme kaydedilirken bir hata oluştu:\n{str(e)}"
                )

    def generate_monthly_report(self):
        month = self.month_combo.currentText()
        year = self.year_spin.value()
        
        with DatabaseConnection() as cursor:
            cursor.execute('''
                SELECT SUM(amount) FROM payments
                WHERE strftime('%Y-%m', payment_date) = ?
            ''', (f"{year}-{self.month_combo.currentIndex()+1:02d}",))
            
            total = cursor.fetchone()[0] or 0
        
        QMessageBox.information(self, "Aylık Rapor",
                              f"{month} {year} Gelir Toplamı: {total} TL")

    def export_to_excel(self, period):
        import_excel_modules()
        try:
            # Kayıt yerini seç
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Dosyasını Kaydet",
                f"odemeler_raporu_{datetime.now().strftime('%Y%m-%d')}.xlsx",
                "Excel Dosyaları (*.xlsx)"
            )
            
            if not file_name:  # Kullanıcı iptal ettiyse
                return
            
            # Excel yazıcı oluştur
            writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
            workbook = writer.book
            
            # Formatları oluştur
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D8E4BC',
                'border': 1
            })
            
            # Verileri çek
            with DatabaseConnection() as cursor:
                cursor.execute('''
                    SELECT 
                        s.name || ' ' || s.surname as student_name,
                        s.phone,
                        s.fee_amount as regular_fee,
                        COUNT(p.id) as payment_count,
                        SUM(p.amount) as total_paid,
                        AVG(p.amount) as avg_payment,
                        MAX(p.payment_date) as last_payment,
                        s.payment_day as due_day,
                        CASE 
                            WHEN s.current_month_paid = 1 THEN 'Ödendi'
                            ELSE 'Ödenmedi'
                        END as current_status,
                        s.blood_type
                    FROM students s
                    LEFT JOIN payments p ON s.id = p.student_id
                    AND p.payment_date BETWEEN ? AND ?
                    GROUP BY s.id
                    ORDER BY student_name
                ''', (datetime.now().strftime("%Y-%m-%d"), datetime.now().strftime("%Y-%m-%d")))
                
                payments = cursor.fetchall()
            
            if not payments:
                QMessageBox.warning(self, "Uyarı", "Bu tarih aralığında ödeme kaydı bulunamadı!")
                return
            
            # DataFrame'leri oluştur
            df = pd.DataFrame(payments, columns=[
                'Öğrenci Adı', 'Telefon', 'Normal Aidat', 'Ödeme Sayısı', 
                'Toplam Ödenen', 'Ortalama Ödeme', 'Son Ödeme Tarihi', 
                'Ödeme Günü', 'Mevcut Durum', 'Kan Grubu'
            ])
            
            # NaN değerleri düzelt
            df = df.fillna({
                'Normal Aidat': 0,
                'Ödeme Sayısı': 0,
                'Toplam Ödenen': 0,
                'Ortalama Ödeme': 0,
                'Son Ödeme Tarihi': '-',
                'Telefon': '-',
                'Kan Grubu': '-'
            })
            
            # Sayısal sütunları float'a çevir
            numeric_columns = ['Normal Aidat', 'Ödeme Sayısı', 'Toplam Ödenen', 'Ortalama Ödeme']
            for col in numeric_columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # Aylık özet için pivot tablo
            pivot_df = pd.pivot_table(
                df,
                values='Ödeme Sayısı',
                index=['Ödeme Günü'],
                aggfunc=['count']
            )
            pivot_df.columns = ['Ödeme Sayısı']
            
            # Başlık sayfası
            summary_df = pd.DataFrame([
                [f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"],
                [f"Başlangıç: {datetime.now().strftime('%d.%m.%Y')}"],
                [f"Bitiş: {datetime.now().strftime('%d.%m.%Y')}"],
                [''],
                [f"Toplam Öğrenci: {len(payments)}"],
                [f"Toplam Tahsilat: {df['Toplam Ödenen'].sum()} TL"]
            ])
            
            # DataFrame'leri Excel'e yaz
            df.to_excel(writer, sheet_name='Ödemeler', index=False)
            pivot_df.to_excel(writer, sheet_name='Aylık Özet')
            summary_df.to_excel(writer, sheet_name='Özet', index=False, header=False)
            
            # Sayfalara eriş
            payments_sheet = writer.sheets['Ödemeler']
            summary_sheet = writer.sheets['Özet']
            monthly_sheet = writer.sheets['Aylık Özet']
            
            # Sütun genişliklerini ayarla
            payments_sheet.set_column('A:A', 30)  # Öğrenci adı
            payments_sheet.set_column('B:B', 15)  # Telefon
            payments_sheet.set_column('C:C', 15)  # Normal Aidat
            payments_sheet.set_column('D:D', 15)  # Ödeme Sayısı
            payments_sheet.set_column('E:E', 15)  # Toplam Ödenen
            payments_sheet.set_column('F:F', 15)  # Ortalama Ödeme
            payments_sheet.set_column('G:G', 15)  # Son Ödeme Tarihi
            payments_sheet.set_column('H:H', 15)  # Ödeme Günü
            payments_sheet.set_column('I:I', 15)  # Mevcut Durum
            payments_sheet.set_column('J:J', 15)  # Kan Grubu
            
            # Başlıkları formatla
            for col_num, value in enumerate(df.columns.values):
                payments_sheet.write(0, col_num, value, header_format)
            
            # Sayısal değerleri formatla
            for row in range(1, len(df) + 1):
                try:
                    payments_sheet.write(row, 2, float(df.iloc[row-1]['Normal Aidat']), header_format)
                    payments_sheet.write(row, 4, float(df.iloc[row-1]['Toplam Ödenen']), header_format)
                    payments_sheet.write(row, 5, float(df.iloc[row-1]['Ortalama Ödeme']), header_format)
                except (ValueError, TypeError):
                    continue
            
            # Özet bilgileri ekle
            summary_row = len(df) + 3  # Boş satır bırak
            
            # Başlık formatı
            title_format = workbook.add_format({
                'bold': True,
                'font_size': 12,
                'bg_color': '#4F81BD',
                'font_color': 'white',
                'align': 'center',
                'border': 1
            })
            
            # Alt toplam formatı
            subtotal_format = workbook.add_format({
                'bold': True,
                'bg_color': '#DCE6F1',
                'border': 1
            })
            
            # Özet başlığı
            payments_sheet.merge_range(summary_row, 0, summary_row, 9, 'ÖZET BİLGİLER', title_format)
            
            # Genel istatistikler
            summary_row += 2
            payments_sheet.write(summary_row, 0, 'Toplam Öğrenci Sayısı:', subtotal_format)
            payments_sheet.write(summary_row, 1, len(df), subtotal_format)
            
            summary_row += 1
            payments_sheet.write(summary_row, 0, 'Toplam Aidat Geliri:', subtotal_format)
            payments_sheet.write(summary_row, 1, float(df['Toplam Ödenen'].sum()), header_format)
            
            summary_row += 1
            payments_sheet.write(summary_row, 0, 'Ortalama Aidat:', subtotal_format)
            payments_sheet.write(summary_row, 1, float(df['Normal Aidat'].mean()), header_format)
            
            summary_row += 1
            payments_sheet.write(summary_row, 0, 'Ortalama Ödeme:', subtotal_format)
            payments_sheet.write(summary_row, 1, float(df['Ortalama Ödeme'].mean()), header_format)
            
            # Ödeme durumu istatistikleri
            summary_row += 2
            payments_sheet.write(summary_row, 0, 'Ödeme Durumu:', subtotal_format)
            
            paid_count = len(df[df['Mevcut Durum'] == 'Ödendi'])
            unpaid_count = len(df[df['Mevcut Durum'] == 'Ödenmedi'])
            
            summary_row += 1
            payments_sheet.write(summary_row, 0, 'Ödeyen Öğrenci Sayısı:', subtotal_format)
            payments_sheet.write(summary_row, 1, paid_count, subtotal_format)
            
            summary_row += 1
            payments_sheet.write(summary_row, 0, 'Ödemeyen Öğrenci Sayısı:', subtotal_format)
            payments_sheet.write(summary_row, 1, unpaid_count, subtotal_format)
            
            summary_row += 1
            payments_sheet.write(summary_row, 0, 'Ödeme Oranı:', subtotal_format)
            payment_rate = (paid_count / len(df) * 100) if len(df) > 0 else 0
            payments_sheet.write(summary_row, 1, f'%{payment_rate:.1f}', subtotal_format)
            
            # Sütun genişliklerini ayarla
            payments_sheet.set_column('A:A', 30)  # Öğrenci adı
            payments_sheet.set_column('B:B', 15)  # Telefon
            payments_sheet.set_column('C:F', 15)  # Sayısal değerler
            payments_sheet.set_column('G:G', 20)  # Tarih
            payments_sheet.set_column('H:J', 15)  # Diğer bilgiler
            
            # Excel dosyasını kaydet ve kapat
            writer.close()
            
            QMessageBox.information(
                self,
                "Başarılı",
                f"Rapor başarıyla kaydedildi:\n{file_name}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Hata",
                f"Rapor oluşturulurken bir hata oluştu:\n{str(e)}"
            )

    def export_to_pdf(self):
        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "PDF Dosyasını Kaydet",
                f"muhasebe_raporu_{datetime.now().strftime('%Y%m-%d')}.pdf",
                "PDF Dosyaları (*.pdf)"
            )
            
            if not file_name:
                return
            
            # PDF dokümanı oluştur
            doc = SimpleDocTemplate(
                file_name,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72,
                encoding='utf-8'  # UTF-8 encoding ekle
            )
            
            # İçerik listesi
            elements = []
            
            # Başlık stili
            styles = getSampleStyleSheet()
            
            # Font kontrolü ve stil tanımlamaları
            try:
                pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))
                font_name = 'DejaVuSans'
            except:
                font_name = 'Helvetica'
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                fontName=font_name,
                encoding='utf-8'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=12,
                fontName=font_name,
                encoding='utf-8'
            )
            
            # Başlık
            title = Paragraph("Muhasebe Raporu", title_style)
            elements.append(title)
            elements.append(Spacer(1, 30))
            
            # Özet Tablo
            summary_data = [
                ['Özet Bilgiler', 'Tutar'],
                ['Toplam Gelir', self.total_label.text()],
                ['Aidat Gelirleri', self.fees_income_label.text()],
                ['Diğer Gelirler', self.other_income_label.text()],
                ['Toplam Gider', self.total_expense_label.text()],
                ['Net Bakiye', self.net_balance_label.text()]
            ]
            
            summary_table = Table(summary_data, colWidths=[300, 200])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Gelirler Tablosu
            elements.append(Paragraph("Gelirler", heading_style))
            elements.append(Spacer(1, 12))
            
            income_data = [['Tarih', 'Tür', 'Açıklama', 'Tutar']]
            for row in range(self.income_table.rowCount()):
                income_data.append([
                    self.income_table.item(row, 0).text(),
                    self.income_table.item(row, 1).text(),
                    self.income_table.item(row, 2).text(),
                    self.income_table.item(row, 3).text()
                ])
            
            income_table = Table(income_data, colWidths=[100, 100, 200, 100])
            income_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(income_table)
            elements.append(Spacer(1, 20))
            
            # Giderler Tablosu
            elements.append(Paragraph("Giderler", heading_style))
            elements.append(Spacer(1, 12))
            
            expense_data = [['Tarih', 'Kategori', 'Açıklama', 'Tutar']]
            for row in range(self.expense_table.rowCount()):
                expense_data.append([
                    self.expense_table.item(row, 0).text(),
                    self.expense_table.item(row, 1).text(),
                    self.expense_table.item(row, 2).text(),
                    self.expense_table.item(row, 3).text()
                ])
            
            expense_table = Table(expense_data, colWidths=[100, 100, 200, 100])
            expense_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(expense_table)
            
            # PDF oluştur
            doc.build(elements)
            
            QMessageBox.information(
                self,
                "Başarılı",
                f"PDF raporu başarıyla kaydedildi:\n{file_name}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Hata",
                f"PDF raporu oluşturulurken bir hata oluştu:\n{str(e)}"
            )

    def aidat_ode(self, ogrenci_id):
        try:
            bugun = datetime.now()
            
            # Eğer 1-7 arası değilse ve yeni kayıt değilse uyarı ver
            if bugun.day > 7:
                QMessageBox.warning(
                    self,
                    "Uyarı",
                    "Aidat ödemeleri her ayın 1-7'si arasında yapılmalıdır!"
                )
                return
                
            # Ödeme durumunu kontrol et
            if kontrol_odeme_durumu(ogrenci_id):
                QMessageBox.warning(self, "Uyarı", "Bu ay için ödeme zaten yapılmış!")
                return
            
            # Öğrencinin aidat tutarını al
            conn = sqlite3.connect('korfez_spor.db')
            cursor = conn.cursor()
            cursor.execute("SELECT fee FROM students WHERE id = ?", (ogrenci_id,))
            aidat_tutari = cursor.fetchone()[0]
            
            # Ödemeyi kaydet
            cursor.execute("""
                INSERT INTO payments 
                (student_id, amount, payment_date, payment_month, payment_year, status) 
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                ogrenci_id,
                aidat_tutari,
                bugun.strftime('%Y-%m-%d'),
                bugun.month,
                bugun.year,
                'Ödendi'
            ))
            
            conn.commit()
            conn.close()
            
            self.load_payments()  # Tabloyu güncelle
            QMessageBox.information(self, "Başarılı", "Ödeme kaydedildi!")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ödeme kaydedilirken hata oluştu: {str(e)}")

class NotesTab(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Üst panel
        top_panel = QHBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Not ara...")
        top_panel.addWidget(QLabel("Ara:"))
        top_panel.addWidget(self.search_input)
        
        add_btn = QPushButton("Not Ekle")
        delete_btn = QPushButton("Sil")
        
        top_panel.addWidget(add_btn)
        top_panel.addWidget(delete_btn)
        top_panel.addStretch()
        
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(
            ["ID", "Öğrenci", "Not", "Tarih"]
        )
        
        layout.addLayout(top_panel)
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        
        add_btn.clicked.connect(self.add_note)
        delete_btn.clicked.connect(self.delete_note)
        self.search_input.textChanged.connect(self.search_notes)
        
        self.load_notes()

    def load_notes(self):
        self.table.setRowCount(0)
        with DatabaseConnection() as cursor:
            cursor.execute('''
                SELECT n.id, s.name, s.surname, n.note, n.date
                FROM notes n
                JOIN students s ON n.student_id = s.id
                ORDER BY n.date DESC
            ''')
            notes = cursor.fetchall()
            
            for note in notes:
                row = self.table.rowCount()
                self.table.insertRow(row)
                
                self.table.setItem(row, 0, QTableWidgetItem(str(note[0])))
                self.table.setItem(row, 1, QTableWidgetItem(f"{note[1]} {note[2]}"))
                self.table.setItem(row, 2, QTableWidgetItem(note[3]))
                self.table.setItem(row, 3, QTableWidgetItem(note[4]))

    def add_note(self):
        dialog = NoteDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            with DatabaseConnection() as cursor:
                cursor.execute('''
                    INSERT INTO notes (student_id, note, date)
                    VALUES (?, ?, ?)
                ''', (
                    dialog.student_combo.currentData(),
                    dialog.note_input.toPlainText(),
                    datetime.now().strftime("%Y-%m-%d %H:%M")
                ))
            self.parent.refresh_all_tabs()

    def delete_note(self):
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen bir not seçin")
            return
        
        reply = QMessageBox.question(self, "Onay", 
                                   "Bu notu silmek istediğinizden emin misiniz?",
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            note_id = self.table.item(current_row, 0).text()
            with DatabaseConnection() as cursor:
                cursor.execute('DELETE FROM notes WHERE id = ?', (note_id,))
            self.parent.refresh_all_tabs()

    def search_notes(self, text):
        for row in range(self.table.rowCount()):
            show = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and text.lower() in item.text().lower():
                    show = True
                    break
            self.table.setRowHidden(row, not show)

class NoteDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Not Ekle")
        self.setModal(True)
        
        layout = QFormLayout()
        
        # Öğrenci seçimi
        self.student_combo = QComboBox()
        self.students = []  # Öğrenci listesi için boş liste oluştur
        self.load_students()  # Öğrencileri yükle
        
        # Not metni
        self.note_text = QTextEdit()
        
        # Tarih
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        
        # Ödeme durumu etiketi
        self.payment_label = QLabel()
        
        layout.addRow("Öğrenci:", self.student_combo)
        layout.addRow("Not:", self.note_text)
        layout.addRow("Tarih:", self.date_input)
        layout.addRow("Ödeme Durumu:", self.payment_label)
        
        # Butonlar
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        
        layout.addRow(buttons)
        
        self.setLayout(layout)
        
        # Öğrenci değiştiğinde ödeme bilgisini güncelle
        self.student_combo.currentIndexChanged.connect(self.update_payment_info)
        
        # İlk öğrenci için ödeme bilgisini göster
        self.update_payment_info()
    
    def load_students(self):
        """Öğrenci listesini yükle"""
        self.student_combo.clear()
        self.students.clear()
        
        with DatabaseConnection() as cursor:
            cursor.execute("""
                SELECT id, name || ' ' || surname, payment_status 
                FROM students 
                ORDER BY name, surname
            """)
            
            for student in cursor.fetchall():
                self.students.append({
                    'id': student[0],
                    'name': student[1],
                    'payment_status': student[2]
                })
                self.student_combo.addItem(student[1], student[0])
    
    def update_payment_info(self):
        """Seçili öğrencinin ödeme durumunu güncelle"""
        if self.student_combo.count() > 0 and len(self.students) > 0:
            current_student = self.students[self.student_combo.currentIndex()]
            
            # Ödeme durumuna göre renk ve metin ayarla
            if current_student['payment_status'] == 'Ödendi':
                self.payment_label.setStyleSheet("color: green;")
                self.payment_label.setText("✓ Ödendi")
            else:
                self.payment_label.setStyleSheet("color: red;")
                self.payment_label.setText("✗ Ödenmedi")

class BulkFeeUpdateDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Toplu Aidat Güncelleme")
        self.setModal(True)
        self.setMinimumWidth(600)  # Genişliği artırdık
        
        layout = QVBoxLayout()
        
        # Mevcut aidatları göster
        current_fees_group = QGroupBox("Mevcut Aidatlar")
        current_fees_layout = QVBoxLayout()
        self.fees_table = QTableWidget()
        self.fees_table.setColumnCount(3)
        self.fees_table.setHorizontalHeaderLabels(["Aidat Miktarı", "Öğrenci Sayısı", "Öğrenciler"])
        current_fees_layout.addWidget(self.fees_table)
        current_fees_group.setLayout(current_fees_layout)
        layout.addWidget(current_fees_group)
        
        # Yeni aidat girişi
        update_group = QGroupBox("Aidat Güncelleme")
        update_layout = QFormLayout()
        
        # Tüm öğrenciler için güncelleme
        all_students_group = QGroupBox("Tüm Öğrenciler İçin")
        all_students_layout = QFormLayout()
        
        self.all_fee_input = QLineEdit()
        self.all_fee_input.setPlaceholderText("Yeni aidat miktarı")
        self.all_fee_input.setValidator(QDoubleValidator(0, 100000, 2))
        
        self.all_fee_timing = QComboBox()
        self.all_fee_timing.addItems(["Bu aydan itibaren", "Bir sonraki aydan itibaren"])
        
        all_students_layout.addRow("Yeni Aidat (TL):", self.all_fee_input)
        all_students_layout.addRow("Uygulama Zamanı:", self.all_fee_timing)
        
        update_all_btn = QPushButton("Tümünü Güncelle")
        update_all_btn.clicked.connect(self.update_all_fees)
        all_students_layout.addRow(update_all_btn)
        
        all_students_group.setLayout(all_students_layout)
        update_layout.addRow(all_students_group)
        
        # Seçili miktar için güncelleme
        selected_group = QGroupBox("Seçili Miktar İçin")
        selected_layout = QFormLayout()
        
        self.selected_fee_input = QLineEdit()
        self.selected_fee_input.setPlaceholderText("Yeni aidat miktarı")
        self.selected_fee_input.setValidator(QDoubleValidator(0, 100000, 2))
        
        self.selected_fee_timing = QComboBox()
        self.selected_fee_timing.addItems(["Bu aydan itibaren", "Bir sonraki aydan itibaren"])
        
        selected_layout.addRow("Yeni Aidat (TL):", self.selected_fee_input)
        selected_layout.addRow("Uygulama Zamanı:", self.selected_fee_timing)
        
        update_selected_btn = QPushButton("Seçili Miktarı Güncelle")
        update_selected_btn.clicked.connect(self.update_selected_fee)
        selected_layout.addRow(update_selected_btn)
        
        selected_group.setLayout(selected_layout)
        update_layout.addRow(selected_group)
        
        update_group = QGroupBox("Aidat Güncelleme")
        update_layout.addRow(update_group)
        layout.addLayout(update_layout)
        
        # İptal butonu
        cancel_btn = QPushButton("İptal")
        cancel_btn.clicked.connect(self.reject)
        layout.addWidget(cancel_btn)
        
        self.setLayout(layout)
        
        # Mevcut aidatları yükle
        self.load_current_fees()
    
    def load_current_fees(self):
        self.fees_table.setRowCount(0)
        fee_groups = {}
        
        with DatabaseConnection() as cursor:
            # fee_amount yerine fee kullan
            cursor.execute('SELECT fee, name, surname FROM students')
            students = cursor.fetchall()
            
            for student in students:
                fee = float(student[0]) if student[0] else 0.0  # None kontrolü ekle
                if fee not in fee_groups:
                    fee_groups[fee] = {
                        'count': 0,
                        'students': []
                    }
                fee_groups[fee]['count'] += 1
                fee_groups[fee]['students'].append(f"{student[1]} {student[2]}")
        
        for row, (fee, info) in enumerate(sorted(fee_groups.items())):
            self.fees_table.insertRow(row)
            self.fees_table.setItem(row, 0, QTableWidgetItem(f"{fee} TL"))
            self.fees_table.setItem(row, 1, QTableWidgetItem(str(info['count'])))
            self.fees_table.setItem(row, 2, QTableWidgetItem(", ".join(info['students'])))
        
        self.fees_table.resizeColumnsToContents()
    
    def update_all_fees(self):
        try:
            new_fee = float(self.all_fee_input.text())
            if new_fee <= 0:
                raise ValueError()
            
            is_current_month = self.all_fee_timing.currentIndex() == 0
            month_text = "bu ay" if is_current_month else "gelecek ay"
            
            reply = QMessageBox.question(
                self,
                "Onay",
                f"Tüm sporcuların aidatı {month_text}dan itibaren {new_fee} TL olarak güncellenecek. Onaylıyor musunuz?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                with DatabaseConnection() as cursor:
                    # fee_amount yerine fee kullan
                    cursor.execute('UPDATE students SET fee = ?', (new_fee,))
                    QMessageBox.information(self, "Başarılı", f"Tüm aidatlar {month_text}dan itibaren güncellendi!")
                    self.accept()
                    
        except ValueError:
            QMessageBox.warning(self, "Hata", "Lütfen geçerli bir aidat miktarı girin!")
    
    def update_selected_fee(self):
        try:
            current_row = self.fees_table.currentRow()
            if current_row < 0:
                QMessageBox.warning(self, "Uyarı", "Lütfen bir aidat miktarı seçin!")
                return
            
            current_fee = float(self.fees_table.item(current_row, 0).text().replace(" TL", ""))
            new_fee = float(self.selected_fee_input.text())
            
            if new_fee <= 0:
                raise ValueError()
            
            is_current_month = self.selected_fee_timing.currentIndex() == 0
            month_text = "bu ay" if is_current_month else "gelecek ay"
            
            reply = QMessageBox.question(
                self,
                "Onay",
                f"{current_fee} TL aidat ödeyen sporcuların aidatı {month_text}dan itibaren {new_fee} TL olarak güncellenecek. Onaylıyor musunuz?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                with DatabaseConnection() as cursor:
                    # fee_amount yerine fee kullan
                    cursor.execute('UPDATE students SET fee = ? WHERE fee = ?', (new_fee, current_fee))
                    QMessageBox.information(self, "Başarılı", f"Seçili aidatlar {month_text}dan itibaren güncellendi!")
                    self.accept()
                    
        except ValueError:
            QMessageBox.warning(self, "Hata", "Lütfen geçerli bir aidat miktarı girin!")

class AgeGroupManagerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Grup Yönetimi")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        
        # Grup listesi
        self.group_table = QTableWidget()
        self.group_table.setColumnCount(3)
        self.group_table.setHorizontalHeaderLabels(["ID", "Grup Adı", "Açıklama"])
        self.group_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.group_table.setSelectionMode(QTableWidget.SingleSelection)
        
        # Yeni grup ekleme formu
        form_group = QGroupBox("Yeni Grup Ekle")
        form_layout = QFormLayout()
        
        self.name_input = QLineEdit()
        self.description_input = QTextEdit()
        self.description_input.setMaximumHeight(60)
        
        form_layout.addRow("Grup Adı:", self.name_input)
        form_layout.addRow("Açıklama:", self.description_input)
        
        form_group.setLayout(form_layout)
        
        # Butonlar
        button_layout = QHBoxLayout()
        add_btn = QPushButton("Ekle")
        delete_btn = QPushButton("Sil")
        close_btn = QPushButton("Kapat")
        
        button_layout.addWidget(add_btn)
        button_layout.addWidget(delete_btn)
        button_layout.addWidget(close_btn)
        
        # Layout'a ekle
        layout.addWidget(self.group_table)
        layout.addWidget(form_group)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Bağlantılar
        add_btn.clicked.connect(self.add_group)
        delete_btn.clicked.connect(self.delete_group)
        close_btn.clicked.connect(self.accept)
        
        # Grupları yükle
        self.load_groups()
    
    def load_groups(self):
        try:
            self.group_table.setRowCount(0)
            with DatabaseConnection() as cursor:
                cursor.execute('SELECT id, name, description FROM groups ORDER BY name')
                groups = cursor.fetchall()
                
                for group in groups:
                    row = self.group_table.rowCount()
                    self.group_table.insertRow(row)
                    
                    for col, value in enumerate(group):
                        self.group_table.setItem(row, col, QTableWidgetItem(str(value)))
                        
            # Sütun genişliklerini ayarla
            self.group_table.setColumnWidth(0, 50)   # ID
            self.group_table.setColumnWidth(1, 150)  # Grup Adı
            self.group_table.setColumnWidth(2, 200)  # Açıklama
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Gruplar yüklenirken hata oluştu: {str(e)}")
    
    def add_group(self):
        try:
            name = self.name_input.text().strip()
            description = self.description_input.toPlainText().strip()
            
            if not name:
                QMessageBox.warning(self, "Uyarı", "Grup adı boş olamaz!")
                return
            
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    INSERT INTO groups (name, description)
                    VALUES (?, ?)
                """, (name, description))
            
            self.name_input.clear()
            self.description_input.clear()
            self.load_groups()
            
            QMessageBox.information(self, "Başarılı", "Grup başarıyla eklendi.")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Grup eklenirken hata oluştu: {str(e)}")
    
    def delete_group(self):
        try:
            current_row = self.group_table.currentRow()
            if current_row < 0:
                QMessageBox.warning(self, "Uyarı", "Lütfen bir grup seçin!")
                return
            
            group_id = self.group_table.item(current_row, 0).text()
            group_name = self.group_table.item(current_row, 1).text()
            
            reply = QMessageBox.question(
                self,
                "Onay",
                f"{group_name} grubunu silmek istediğinizden emin misiniz?\n\n"
                "Not: Bu gruba ait öğrenciler grupsuz kalacaktır.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                with DatabaseConnection() as cursor:
                    # Önce öğrencilerin grup_id'sini NULL yap
                    cursor.execute("UPDATE students SET group_id = NULL WHERE group_id = ?", (group_id,))
                    # Sonra grubu sil
                    cursor.execute("DELETE FROM groups WHERE id = ?", (group_id,))
                
                self.load_groups()
                QMessageBox.information(self, "Başarılı", "Grup başarıyla silindi.")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Grup silinirken hata oluştu: {str(e)}")

class GroupDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Yaş Grubu")
        self.setModal(True)
        
        layout = QFormLayout()
        
        self.name_input = QLineEdit()
        self.description_input = QTextEdit()
        self.description_input.setMaximumHeight(100)
        
        layout.addRow("Grup Adı:", self.name_input)
        layout.addRow("Açıklama:", self.description_input)
        
        buttons = QHBoxLayout()
        save_btn = QPushButton("Kaydet")
        cancel_btn = QPushButton("İptal")
        
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        
        buttons.addWidget(save_btn)
        buttons.addWidget(cancel_btn)
        
        layout.addRow(buttons)
        self.setLayout(layout)

class UnpaidStudentsTab(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Üst panel
        top_panel = QHBoxLayout()
        
        # Ara alanı
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Sporcu ara...")
        top_panel.addWidget(QLabel("Ara:"))
        top_panel.addWidget(self.search_input)
        
        # Tablo
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Öğrenci", "Telefon", "Aidat Günü", "Son Ödeme", "Gecikme", "Aidat"
        ])
        
        # Sütun genişliklerini ayarla
        self.table.setColumnWidth(0, 200)  # Öğrenci
        self.table.setColumnWidth(1, 120)  # Telefon
        self.table.setColumnWidth(2, 80)   # Aidat Günü
        self.table.setColumnWidth(3, 100)  # Son Ödeme
        self.table.setColumnWidth(4, 100)  # Gecikme
        self.table.setColumnWidth(5, 100)  # Aidat
        
        layout.addLayout(top_panel)
        layout.addWidget(self.table)
        
        # Özet bilgileri
        summary_panel = QHBoxLayout()
        self.total_unpaid_label = QLabel("Toplam Ödeme Bekleyen: 0")
        self.total_amount_label = QLabel("Toplam Beklenen Tutar: 0 TL")
        
        summary_panel.addWidget(self.total_unpaid_label)
        summary_panel.addWidget(self.total_amount_label)
        summary_panel.addStretch()
        
        layout.addLayout(summary_panel)
        
        self.setLayout(layout)
        
        # Bağlantılar
        self.search_input.textChanged.connect(self.search_students)
        
        self.load_unpaid_students()
    
    def load_unpaid_students(self):
        try:
            self.table.setRowCount(0)
            current_month = datetime.now().month
            current_year = datetime.now().year
            
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT 
                        s.id,
                        s.name || ' ' || s.surname as student_name,
                        s.payment_day,
                        MAX(p.payment_date) as last_payment,
                        s.fee,
                        s.payment_status,
                        g.name as group_name
                    FROM students s
                    LEFT JOIN payments p ON s.id = p.student_id
                    LEFT JOIN groups g ON s.group_id = g.id
                    WHERE s.payment_status = 'Ödenmedi'
                        OR s.payment_status = 'GECİKMİŞ ÖDEME'
                    GROUP BY s.id
                    ORDER BY s.name
                """)
                
                students = cursor.fetchall()
                
                for student in students:
                    row = self.table.rowCount()
                    self.table.insertRow(row)
                    
                    # Öğrenci ID
                    self.table.setItem(row, 0, QTableWidgetItem(str(student[0])))
                    
                    # Öğrenci Adı
                    self.table.setItem(row, 1, QTableWidgetItem(student[1]))
                    
                    # Aidat Günü
                    self.table.setItem(row, 2, QTableWidgetItem(str(student[2])))
                    
                    # Son Ödeme
                    last_payment = student[3] or "Ödeme yapılmamış"
                    self.table.setItem(row, 3, QTableWidgetItem(str(last_payment)))
                    
                    # Aidat Tutarı
                    self.table.setItem(row, 4, QTableWidgetItem(f"{student[4]} TL"))
                    
                    # Ödeme Durumu
                    status_item = QTableWidgetItem(student[5])
                    status_item.setForeground(Qt.red)
                    self.table.setItem(row, 5, status_item)
                    
                    # Grup
                    self.table.setItem(row, 6, QTableWidgetItem(student[6] or ""))
            
            self.total_unpaid_label.setText(f"Toplam Ödeme Bekleyen: {len(students)}")
            self.total_amount_label.setText(f"Toplam Beklenen Tutar: {sum(student[4] for student in students)} TL")
            
            # Tabloyu gecikmeye göre sırala
            self.table.sortItems(4, Qt.DescendingOrder)
        
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ödenmemiş öğrenciler yüklenirken hata oluştu: {str(e)}")

    def search_students(self, text):
        text = text.lower()
        for row in range(self.table.rowCount()):
            show = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and text in item.text().lower():
                    show = True
                    break
            self.table.setRowHidden(row, not show)

class AccountingTab(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Üst panel - Tarih Filtreleme
        filter_panel = QHBoxLayout()
        
        # Tarih aralığı seçimi
        self.date_filter = QComboBox()
        self.date_filter.addItems([
            "Bu Ay", "Geçen Ay", "Son 3 Ay", "Son 6 Ay", 
            "Bu Yıl", "Geçen Yıl", "Tüm Zamanlar", "Özel Aralık"
        ])
        
        self.start_date = QDateEdit()
        self.end_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addMonths(-1))
        self.end_date.setDate(QDate.currentDate())
        self.start_date.hide()
        self.end_date.hide()
        
        filter_panel.addWidget(QLabel("Tarih Aralığı:"))
        filter_panel.addWidget(self.date_filter)
        filter_panel.addWidget(self.start_date)
        filter_panel.addWidget(QLabel("-"))
        filter_panel.addWidget(self.end_date)
        
        # Yenile butonu ekle
        refresh_btn = QPushButton("Yenile")
        refresh_btn.setIcon(QIcon.fromTheme("view-refresh"))  # Sistem simgesi kullan
        refresh_btn.clicked.connect(self.load_data)  # Yenileme fonksiyonuna bağla
        filter_panel.addWidget(refresh_btn)
        
        # Excel çıktı butonu
        excel_btn = QPushButton("Excel'e Aktar")
        excel_btn.clicked.connect(self.export_to_excel)  # Bağlantı eklendi
        filter_panel.addWidget(excel_btn)
        
        # PDF çıktı butonu
        pdf_btn = QPushButton("PDF'e Aktar")
        pdf_btn.clicked.connect(self.export_to_pdf)  # Bağlantı eklendi
        filter_panel.addWidget(pdf_btn)
        
        filter_panel.addStretch()
        
        layout.addLayout(filter_panel)
        
        # Özet Panel
        summary_group = QGroupBox("Özet Bilgiler")
        summary_layout = QGridLayout()
        
        self.total_income_label = QLabel("0.00 TL")
        self.total_expense_label = QLabel("0.00 TL")
        self.net_balance_label = QLabel("0.00 TL")
        self.fees_income_label = QLabel("0.00 TL")
        self.other_income_label = QLabel("0.00 TL")
        
        summary_layout.addWidget(QLabel("Toplam Gelir:"), 0, 0)
        summary_layout.addWidget(self.total_income_label, 0, 1)
        summary_layout.addWidget(QLabel("Aidat Gelirleri:"), 0, 2)
        summary_layout.addWidget(self.fees_income_label, 0, 3)
        summary_layout.addWidget(QLabel("Diğer Gelirler:"), 0, 4)
        summary_layout.addWidget(self.other_income_label, 0, 5)
        
        summary_layout.addWidget(QLabel("Toplam Gider:"), 1, 0)
        summary_layout.addWidget(self.total_expense_label, 1, 1)
        summary_layout.addWidget(QLabel("Net Bakiye:"), 1, 2)
        summary_layout.addWidget(self.net_balance_label, 1, 3)
        
        summary_group.setLayout(summary_layout)
        layout.addWidget(summary_group)
        
        # Gelir/Gider Tabloları
        tables_layout = QHBoxLayout()
        
        # Gelirler Tablosu
        income_group = QGroupBox("Gelirler")
        income_layout = QVBoxLayout()
        
        add_income_btn = QPushButton("Gelir Ekle")
        income_layout.addWidget(add_income_btn)
        
        self.income_table = QTableWidget()
        self.income_table.setColumnCount(4)
        self.income_table.setHorizontalHeaderLabels(["Tarih", "Tür", "Açıklama", "Tutar"])
        
        # Gelir butonları
        income_buttons = QHBoxLayout()
        delete_income_btn = QPushButton("Seçili Geliri Sil")  # Yeni buton
        
        income_buttons.addWidget(delete_income_btn)  # Yeni buton eklendi
        income_buttons.addStretch()
        
        income_layout.addWidget(self.income_table)
        income_layout.addLayout(income_buttons)
        income_group.setLayout(income_layout)
        
        tables_layout.addWidget(income_group)
        
        # Giderler Tablosu
        expense_group = QGroupBox("Giderler")
        expense_layout = QVBoxLayout()
        
        add_expense_btn = QPushButton("Gider Ekle")
        expense_layout.addWidget(add_expense_btn)
        
        self.expense_table = QTableWidget()
        self.expense_table.setColumnCount(4)
        self.expense_table.setHorizontalHeaderLabels(["Tarih", "Kategori", "Açıklama", "Tutar"])
        
        # Gider butonları
        expense_buttons = QHBoxLayout()
        delete_expense_btn = QPushButton("Seçili Gideri Sil")  # Yeni buton
        
        expense_buttons.addWidget(delete_expense_btn)  # Yeni buton eklendi
        expense_buttons.addStretch()
        
        expense_layout.addWidget(self.expense_table)
        expense_layout.addLayout(expense_buttons)
        expense_group.setLayout(expense_layout)
        
        tables_layout.addWidget(expense_group)
        
        layout.addLayout(tables_layout)
        
        self.setLayout(layout)
        
        # Bağlantılar
        self.date_filter.currentIndexChanged.connect(self.handle_date_filter_change)
        add_income_btn.clicked.connect(self.add_income)
        delete_income_btn.clicked.connect(self.delete_income)  # Yeni bağlantı
        add_expense_btn.clicked.connect(self.add_expense)
        delete_expense_btn.clicked.connect(self.delete_expense)  # Yeni bağlantı
        
        # Veritabanı tablosunu oluştur
        self.create_tables()
        
        # Verileri yükle
        self.load_data()
    
    def create_tables(self):
        with DatabaseConnection() as cursor:
            # Gelirler tablosu
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS incomes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    type TEXT NOT NULL,
                    description TEXT,
                    amount REAL NOT NULL
                )
            ''')
            
            # Giderler tablosu
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS expenses (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    category TEXT NOT NULL,
                    description TEXT,
                    amount REAL NOT NULL
                )
            ''')
    
    def handle_date_filter_change(self, index):
        show_custom = index == 7  # "Özel Aralık" seçeneği
        self.start_date.setVisible(show_custom)
        self.end_date.setVisible(show_custom)
        self.load_data()
    
    def get_date_range(self):
        filter_index = self.date_filter.currentIndex()
        today = QDate.currentDate()
        
        if filter_index == 0:  # Bu Ay
            start = today.addDays(-(today.day() - 1))
            end = today
        elif filter_index == 1:  # Geçen Ay
            start = today.addMonths(-1).addDays(-(today.day() - 1))
            end = today.addDays(-today.day())
        elif filter_index == 2:  # Son 3 Ay
            start = today.addMonths(-3)
            end = today
        elif filter_index == 3:  # Son 6 Ay
            start = today.addMonths(-6)
            end = today
        elif filter_index == 4:  # Bu Yıl
            start = QDate(today.year(), 1, 1)
            end = today
        elif filter_index == 5:  # Geçen Yıl
            start = QDate(today.year() - 1, 1, 1)
            end = QDate(today.year() - 1, 12, 31)
        elif filter_index == 6:  # Tüm Zamanlar
            return None, None
        else:  # Özel Aralık
            start = self.start_date.date()
            end = self.end_date.date()
        
        return start.toString("yyyy-MM-dd"), end.toString("yyyy-MM-dd")
    
    def load_data(self):
        try:
            start_date, end_date = self.get_date_range()
            
            # Tabloları temizle
            self.income_table.setRowCount(0)
            self.expense_table.setRowCount(0)
            
            total_fees = 0
            total_other_income = 0
            total_expenses = 0
            
            with DatabaseConnection() as cursor:
                # Sadece payments tablosundan aidat gelirlerini al
                query = '''
                    SELECT 
                        p.payment_date,
                        s.name || ' ' || s.surname as student_name,
                        p.amount,
                        p.payment_month,
                        p.payment_year
                    FROM payments p
                    JOIN students s ON p.student_id = s.id
                    WHERE p.status = 'Ödendi'
                '''
                
                if start_date and end_date:
                    query += ' AND p.payment_date BETWEEN ? AND ?'
                    cursor.execute(query, (start_date, end_date))
                else:
                    cursor.execute(query)
                
                # Aidat ödemelerini tabloya ekle
                for payment in cursor.fetchall():
                    row = self.income_table.rowCount()
                    self.income_table.insertRow(row)
                    
                    payment_date = datetime.strptime(payment[0], '%Y-%m-%d')
                    
                    self.income_table.setItem(row, 0, QTableWidgetItem(payment_date.strftime('%d/%m/%Y')))
                    self.income_table.setItem(row, 1, QTableWidgetItem("Aidat"))
                    self.income_table.setItem(row, 2, QTableWidgetItem(
                        f"{payment[1]} - {payment[3]}/{payment[4]} ayı aidatı"
                    ))
                    self.income_table.setItem(row, 3, QTableWidgetItem(f"{payment[2]:.2f} TL"))
                    
                    total_fees += payment[2]
                
                # Diğer gelirleri income tablosundan al (aidat dışındaki gelirler)
                query = '''
                    SELECT date, type, description, amount 
                    FROM income 
                    WHERE type != 'Aidat'
                '''
                
                if start_date and end_date:
                    query += ' AND date BETWEEN ? AND ?'
                    cursor.execute(query, (start_date, end_date))
                else:
                    cursor.execute(query)
                
                # Diğer gelirleri tabloya ekle
                for income in cursor.fetchall():
                    row = self.income_table.rowCount()
                    self.income_table.insertRow(row)
                    
                    income_date = datetime.strptime(income[0], '%Y-%m-%d')
                    
                    self.income_table.setItem(row, 0, QTableWidgetItem(income_date.strftime('%d/%m/%Y')))
                    self.income_table.setItem(row, 1, QTableWidgetItem(income[1]))
                    self.income_table.setItem(row, 2, QTableWidgetItem(income[2]))
                    self.income_table.setItem(row, 3, QTableWidgetItem(f"{income[3]:.2f} TL"))
                    
                    total_other_income += income[3]
                
                # Giderleri yükle
                query = 'SELECT date, category, description, amount FROM expenses'
                if start_date and end_date:
                    query += ' WHERE date BETWEEN ? AND ?'
                    cursor.execute(query, (start_date, end_date))
                else:
                    cursor.execute(query)
                
                # Giderleri tabloya ekle
                for expense in cursor.fetchall():
                    row = self.expense_table.rowCount()
                    self.expense_table.insertRow(row)
                    
                    expense_date = datetime.strptime(expense[0], '%Y-%m-%d')
                    
                    self.expense_table.setItem(row, 0, QTableWidgetItem(expense_date.strftime('%d/%m/%Y')))
                    self.expense_table.setItem(row, 1, QTableWidgetItem(expense[1]))
                    self.expense_table.setItem(row, 2, QTableWidgetItem(expense[2]))
                    self.expense_table.setItem(row, 3, QTableWidgetItem(f"{expense[3]:.2f} TL"))
                    
                    total_expenses += expense[3]
                
                # Özet bilgileri güncelle
                total_income = total_fees + total_other_income
                net_balance = total_income - total_expenses
                
                self.total_income_label.setText(f"{total_income:.2f} TL")
                self.fees_income_label.setText(f"{total_fees:.2f} TL")
                self.other_income_label.setText(f"{total_other_income:.2f} TL")
                self.total_expense_label.setText(f"{total_expenses:.2f} TL")
                self.net_balance_label.setText(f"{net_balance:.2f} TL")
                
                # Bakiye rengini ayarla
                self.net_balance_label.setStyleSheet(
                    "color: green;" if net_balance >= 0 else "color: red;"
                )
                
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Veriler yüklenirken hata oluştu: {str(e)}")
    
    def add_income(self):
        dialog = IncomeDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            try:
                with DatabaseConnection() as cursor:
                    # Geliri income tablosuna ekle
                    cursor.execute('''
                        INSERT INTO income (date, type, description, amount)
                        VALUES (?, ?, ?, ?)
                    ''', (
                        dialog.date_input.date().toString("yyyy-MM-dd"),
                        dialog.type_combo.currentText(),
                        dialog.description_input.text(),
                        float(dialog.amount_input.text())
                    ))
                
                QMessageBox.information(
                    self,
                    "Başarılı",
                    f"Gelir kaydı başarıyla eklendi!\n\n"
                    f"Tür: {dialog.type_combo.currentText()}\n"
                    f"Tutar: {dialog.amount_input.text()} TL\n"
                    f"Tarih: {dialog.date_input.date().toString('dd/MM/yyyy')}"
                )
                
                self.load_data()  # Tabloyu güncelle
                
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Hata",
                    f"Gelir kaydedilirken bir hata oluştu:\n{str(e)}"
                )
    
    def add_expense(self):
        dialog = ExpenseDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            with DatabaseConnection() as cursor:
                cursor.execute('''
                    INSERT INTO expenses (date, category, description, amount)
                    VALUES (?, ?, ?, ?)
                ''', (
                    dialog.date_input.date().toString("yyyy-MM-dd"),
                    dialog.category_combo.currentText(),
                    dialog.description_input.text(),
                    float(dialog.amount_input.text())
                ))
            self.load_data()
    
    def export_to_excel(self):
        import_excel_modules()
        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Excel Dosyasını Kaydet",
                f"muhasebe_raporu_{datetime.now().strftime('%Y%m-%d')}.xlsx",
                "Excel Dosyaları (*.xlsx)"
            )
            
            if not file_name:
                return
            
            writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
            workbook = writer.book
            
            # Formatlar
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4F81BD',
                'font_color': 'white',
                'border': 1,
                'align': 'center'
            })
            
            subheader_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D8E4BC',
                'border': 1
            })
            
            money_format = workbook.add_format({
                'num_format': '#,##0.00 ₺',
                'align': 'right'
            })
            
            date_format = workbook.add_format({
                'num_format': 'dd/mm/yyyy',
                'align': 'center'
            })
            
            positive_money = workbook.add_format({
                'num_format': '#,##0.00 ₺',
                'align': 'right',
                'font_color': 'green'
            })
            
            negative_money = workbook.add_format({
                'num_format': '#,##0.00 ₺',
                'align': 'right',
                'font_color': 'red'
            })
            
            # Özet sayfası
            summary_sheet = workbook.add_worksheet('Özet')
            
            # Başlık
            title = f"MUHASEBE RAPORU ({datetime.now().strftime('%d/%m/%Y')})"
            summary_sheet.merge_range('A1:E1', title, header_format)
            
            # Tarih aralığı
            start_date, end_date = self.get_date_range()
            if start_date and end_date:
                date_range = f"Tarih Aralığı: {start_date} - {end_date}"
            else:
                date_range = "Tarih Aralığı: Tüm Zamanlar"
            summary_sheet.merge_range('A2:E2', date_range, subheader_format)
            
            # Genel Özet
            summary_sheet.write('A4', 'GENEL ÖZET', header_format)
            summary_sheet.write('A5', 'Toplam Gelir:', subheader_format)
            summary_sheet.write('B5', float(self.total_income_label.text().replace(" TL", "")), positive_money)
            
            summary_sheet.write('A6', 'Aidat Gelirleri:', subheader_format)
            summary_sheet.write('B6', float(self.fees_income_label.text().replace(" TL", "")), positive_money)
            
            summary_sheet.write('A7', 'Diğer Gelirler:', subheader_format)
            summary_sheet.write('B7', float(self.other_income_label.text().replace(" TL", "")), positive_money)
            
            summary_sheet.write('A8', 'Toplam Gider:', subheader_format)
            summary_sheet.write('B8', float(self.total_expense_label.text().replace(" TL", "")), negative_money)
            
            net_balance = float(self.net_balance_label.text().replace(" TL", ""))
            summary_sheet.write('A9', 'Net Bakiye:', subheader_format)
            summary_sheet.write('B9', net_balance, positive_money if net_balance >= 0 else negative_money)
            
            # Gelir Dağılımı
            summary_sheet.write('A11', 'GELİR DAĞILIMI', header_format)
            
            # Gelir türlerine göre pasta grafik
            income_chart = workbook.add_chart({'type': 'pie'})
            income_data = [
                ['Aidat Gelirleri', float(self.fees_income_label.text().replace(" TL", ""))],
                ['Diğer Gelirler', float(self.other_income_label.text().replace(" TL", ""))]
            ]
            summary_sheet.write_column('A12', [row[0] for row in income_data])
            summary_sheet.write_column('B12', [row[1] for row in income_data])
            
            income_chart.add_series({
                'name': 'Gelir Dağılımı',
                'categories': '=Özet!$A$12:$A$13',
                'values': '=Özet!$B$12:$B$13'
            })
            income_chart.set_title({'name': 'Gelir Dağılımı'})
            summary_sheet.insert_chart('D4', income_chart)
            
            # Gider Kategorileri
            summary_sheet.write('A15', 'GİDER KATEGORİLERİ', header_format)
            expense_categories = {}
            for row in range(self.expense_table.rowCount()):
                category = self.expense_table.item(row, 1).text()
                amount = float(self.expense_table.item(row, 3).text().replace(" TL", ""))
                expense_categories[category] = expense_categories.get(category, 0) + amount
            
            row_num = 16
            for category, amount in expense_categories.items():
                summary_sheet.write(f'A{row_num}', category, subheader_format)
                summary_sheet.write(f'B{row_num}', amount, negative_money)
                row_num += 1
            
            # Gider dağılımı pasta grafik
            expense_chart = workbook.add_chart({'type': 'pie'})
            expense_chart.add_series({
                'name': 'Gider Dağılımı',
                'categories': f'=Özet!$A$16:$A${row_num-1}',
                'values': f'=Özet!$B$16:$B${row_num-1}'
            })
            expense_chart.set_title({'name': 'Gider Dağılımı'})
            summary_sheet.insert_chart('D15', expense_chart)
            
            # Detaylı Gelirler sayfası
            income_sheet = workbook.add_worksheet('Gelirler')
            income_headers = ['Tarih', 'Tür', 'Açıklama', 'Tutar']
            
            for col, header in enumerate(income_headers):
                income_sheet.write(0, col, header, header_format)
            
            for row in range(self.income_table.rowCount()):
                income_sheet.write(row + 1, 0, self.income_table.item(row, 0).text(), date_format)
                income_sheet.write(row + 1, 1, self.income_table.item(row, 1).text())
                income_sheet.write(row + 1, 2, self.income_table.item(row, 2).text())
                income_sheet.write(row + 1, 3, float(self.income_table.item(row, 3).text().replace(" TL", "")), money_format)
            
            # Detaylı Giderler sayfası
            expense_sheet = workbook.add_worksheet('Giderler')
            expense_headers = ['Tarih', 'Kategori', 'Açıklama', 'Tutar']
            
            for col, header in enumerate(expense_headers):
                expense_sheet.write(0, col, header, header_format)
            
            for row in range(self.expense_table.rowCount()):
                expense_sheet.write(row + 1, 0, self.expense_table.item(row, 0).text(), date_format)
                expense_sheet.write(row + 1, 1, self.expense_table.item(row, 1).text())
                expense_sheet.write(row + 1, 2, self.expense_table.item(row, 2).text())
                expense_sheet.write(row + 1, 3, float(self.expense_table.item(row, 3).text().replace(" TL", "")), money_format)
            
            # Sütun genişliklerini ayarla
            for sheet in [summary_sheet, income_sheet, expense_sheet]:
                sheet.set_column('A:A', 15)  # Tarih
                sheet.set_column('B:B', 20)  # Tür/Kategori
                sheet.set_column('C:C', 40)  # Açıklama
                sheet.set_column('D:D', 15)  # Tutar
            
            writer.close()
            
            QMessageBox.information(
                self,
                "Başarılı",
                f"Muhasebe raporu başarıyla kaydedildi:\n{file_name}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Hata",
                f"Rapor oluşturulurken bir hata oluştu:\n{str(e)}"
            )

    def export_to_pdf(self):
        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "PDF Dosyasını Kaydet",
                f"muhasebe_raporu_{datetime.now().strftime('%Y%m-%d')}.pdf",
                "PDF Dosyaları (*.pdf)"
            )
            
            if not file_name:
                return
            
            # PDF dokümanı oluştur
            doc = SimpleDocTemplate(
                file_name,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72,
                encoding='utf-8'  # UTF-8 encoding ekle
            )
            
            # İçerik listesi
            elements = []
            
            # Başlık stili
            styles = getSampleStyleSheet()
            
            # Font kontrolü ve stil tanımlamaları
            try:
                pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))
                font_name = 'DejaVuSans'
            except:
                font_name = 'Helvetica'
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                fontName=font_name,
                encoding='utf-8'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=12,
                fontName=font_name,
                encoding='utf-8'
            )
            
            # Başlık
            title = Paragraph("Muhasebe Raporu", title_style)
            elements.append(title)
            elements.append(Spacer(1, 30))
            
            # Özet Tablo
            summary_data = [
                ['Özet Bilgiler', 'Tutar'],
                ['Toplam Gelir', self.total_income_label.text()],
                ['Aidat Gelirleri', self.fees_income_label.text()],
                ['Diğer Gelirler', self.other_income_label.text()],
                ['Toplam Gider', self.total_expense_label.text()],
                ['Net Bakiye', self.net_balance_label.text()]
            ]
            
            summary_table = Table(summary_data, colWidths=[300, 200])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Gelirler Tablosu
            elements.append(Paragraph("Gelirler", heading_style))
            elements.append(Spacer(1, 12))
            
            income_data = [['Tarih', 'Tür', 'Açıklama', 'Tutar']]
            for row in range(self.income_table.rowCount()):
                income_data.append([
                    self.income_table.item(row, 0).text(),
                    self.income_table.item(row, 1).text(),
                    self.income_table.item(row, 2).text(),
                    self.income_table.item(row, 3).text()
                ])
            
            income_table = Table(income_data, colWidths=[100, 100, 200, 100])
            income_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(income_table)
            elements.append(Spacer(1, 20))
            
            # Giderler Tablosu
            elements.append(Paragraph("Giderler", heading_style))
            elements.append(Spacer(1, 12))
            
            expense_data = [['Tarih', 'Kategori', 'Açıklama', 'Tutar']]
            for row in range(self.expense_table.rowCount()):
                expense_data.append([
                    self.expense_table.item(row, 0).text(),
                    self.expense_table.item(row, 1).text(),
                    self.expense_table.item(row, 2).text(),
                    self.expense_table.item(row, 3).text()
                ])
            
            expense_table = Table(expense_data, colWidths=[100, 100, 200, 100])
            expense_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(expense_table)
            
            # PDF oluştur
            doc.build(elements)
            
            QMessageBox.information(
                self,
                "Başarılı",
                f"PDF raporu başarıyla kaydedildi:\n{file_name}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Hata",
                f"PDF raporu oluşturulurken bir hata oluştu:\n{str(e)}"
            )

    def delete_income(self):
        try:
            current_row = self.income_table.currentRow()
            if current_row < 0:
                QMessageBox.warning(self, "Uyarı", "Lütfen silinecek geliri seçin!")
                return
            
            # Seçili gelirin bilgilerini al
            tarih = self.income_table.item(current_row, 0).text()
            tur = self.income_table.item(current_row, 1).text()
            aciklama = self.income_table.item(current_row, 2).text()
            tutar = self.income_table.item(current_row, 3).text().replace(" TL", "")
            
            # Onay al
            reply = QMessageBox.question(
                self,
                "Onay",
                f"Bu gelir kaydını silmek istediğinizden emin misiniz?\n\n"
                f"Tarih: {tarih}\n"
                f"Tür: {tur}\n"
                f"Açıklama: {aciklama}\n"
                f"Tutar: {tutar} TL",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                with DatabaseConnection() as cursor:
                    # Tarihi veritabanı formatına çevir
                    db_tarih = datetime.strptime(tarih, '%d/%m/%Y').strftime('%Y-%m-%d')
                    
                    if tur == "Aidat":
                        # Aidat ödemesini sil
                        cursor.execute("""
                            DELETE FROM payments 
                            WHERE payment_date = ? AND amount = ?
                        """, (db_tarih, float(tutar)))
                        
                        # Öğrencinin ödeme durumunu güncelle
                        cursor.execute("""
                            UPDATE students 
                            SET payment_status = 'Ödenmedi'
                            WHERE id IN (
                                SELECT student_id 
                                FROM payments 
                                WHERE payment_date = ? AND amount = ?
                            )
                        """, (db_tarih, float(tutar)))
                    else:
                        # Normal geliri sil
                        cursor.execute("""
                            DELETE FROM income 
                            WHERE date = ? AND type = ? AND description = ? AND amount = ?
                        """, (db_tarih, tur, aciklama, float(tutar)))
                    
                    if cursor.rowcount > 0:
                        # Tabloları güncelle
                        self.load_data()
                        if tur == "Aidat":
                            self.parent.payments_tab.load_payments()
                            self.parent.unpaid_tab.load_unpaid_students()
                        
                        QMessageBox.information(self, "Başarılı", "Gelir kaydı silindi.")
                    else:
                        QMessageBox.warning(self, "Uyarı", "Gelir kaydı silinemedi!")
                
                self.load_data()  # Tabloları yenile
                QMessageBox.information(self, "Başarılı", "Gider kaydı silindi.")
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Gelir silinirken hata oluştu: {str(e)}")

    def delete_expense(self):
        try:
            current_row = self.expense_table.currentRow()
            if current_row < 0:
                QMessageBox.warning(self, "Uyarı", "Lütfen silinecek gideri seçin!")
                return
            
            # Seçili giderin bilgilerini al
            tarih = self.expense_table.item(current_row, 0).text()
            kategori = self.expense_table.item(current_row, 1).text()
            aciklama = self.expense_table.item(current_row, 2).text()
            tutar = self.expense_table.item(current_row, 3).text().replace(" TL", "")
            
            # Onay al
            reply = QMessageBox.question(
                self,
                "Onay",
                f"Bu gider kaydını silmek istediğinizden emin misiniz?\n\n"
                f"Tarih: {tarih}\n"
                f"Kategori: {kategori}\n"
                f"Açıklama: {aciklama}\n"
                f"Tutar: {tutar} TL",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                with DatabaseConnection() as cursor:
                    # Tarihi veritabanı formatına çevir
                    db_tarih = datetime.strptime(tarih, '%d/%m/%Y').strftime('%Y-%m-%d')
                    
                    cursor.execute("""
                        DELETE FROM expenses 
                        WHERE date = ? AND category = ? AND description = ? AND amount = ?
                    """, (db_tarih, kategori, aciklama, float(tutar)))
                    
                    if cursor.rowcount > 0:
                        # Tabloları güncelle
                        self.load_data()
                        QMessageBox.information(self, "Başarılı", "Gider kaydı silindi.")
                    else:
                        QMessageBox.warning(self, "Uyarı", "Gider kaydı silinemedi!")
        
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Gider silinirken hata oluştu: {str(e)}")

class IncomeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Gelir Ekle")
        self.setModal(True)
        
        layout = QFormLayout()
        
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        
        self.type_combo = QComboBox()
        self.type_combo.addItems([
            "Bağış", "Sponsorluk", "Etkinlik Geliri", "Diğer"
        ])
        
        self.description_input = QLineEdit()
        self.amount_input = QLineEdit()
        self.amount_input.setValidator(QDoubleValidator(0, 1000000, 2))
        
        layout.addRow("Tarih:", self.date_input)
        layout.addRow("Tür:", self.type_combo)
        layout.addRow("Açıklama:", self.description_input)
        layout.addRow("Tutar (TL):", self.amount_input)
        
        buttons = QHBoxLayout()
        save_btn = QPushButton("Kaydet")
        cancel_btn = QPushButton("İptal")
        
        save_btn.clicked.connect(self.validate_and_accept)
        cancel_btn.clicked.connect(self.reject)
        
        buttons.addWidget(save_btn)
        buttons.addWidget(cancel_btn)
        
        layout.addRow(buttons)
        self.setLayout(layout)
    
    def validate_and_accept(self):
        if not self.amount_input.text():
            QMessageBox.warning(self, "Hata", "Lütfen tutar girin!")
            return
        
        try:
            amount = float(self.amount_input.text())
            if amount <= 0:
                raise ValueError()
        except ValueError:
            QMessageBox.warning(self, "Hata", "Geçerli bir tutar girin!")
            return
        
        self.accept()

class ExpenseDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Gider Ekle")
        self.setModal(True)
        
        layout = QFormLayout()
        
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        
        self.category_combo = QComboBox()
        self.category_combo.addItems([
            "Saha Kirası", 
            "Personel Maaşı", "Malzeme", "Bakım-Onarım", "Diğer"
        ])
        
        self.description_input = QLineEdit()
        self.amount_input = QLineEdit()
        self.amount_input.setValidator(QDoubleValidator(0, 1000000, 2))
        
        layout.addRow("Tarih:", self.date_input)
        layout.addRow("Kategori:", self.category_combo)
        layout.addRow("Açıklama:", self.description_input)
        layout.addRow("Tutar (TL):", self.amount_input)
        
        buttons = QHBoxLayout()
        save_btn = QPushButton("Kaydet")
        cancel_btn = QPushButton("İptal")
        
        save_btn.clicked.connect(self.validate_and_accept)
        cancel_btn.clicked.connect(self.reject)
        
        buttons.addWidget(save_btn)
        buttons.addWidget(cancel_btn)
        
        layout.addRow(buttons)
        self.setLayout(layout)
    
    def validate_and_accept(self):
        if not self.amount_input.text():
            QMessageBox.warning(self, "Hata", "Lütfen tutar girin!")
            return
        
        try:
            amount = float(self.amount_input.text())
            if amount <= 0:
                raise ValueError()
        except ValueError:
            QMessageBox.warning(self, "Hata", "Geçerli bir tutar girin!")
            return
        
        self.accept()

class PaymentDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Aidat Ödeme")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        
        # Sporcu seçim grubu
        athlete_group = QGroupBox("Sporcu Seçimi")
        athlete_layout = QVBoxLayout()
        
        # Arama kutusu ve otomatik tamamlama
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Sporcu adını yazın...")
        
        # Completer için model
        self.completer = QCompleter()
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.completer.setFilterMode(Qt.MatchContains)
        self.search_input.setCompleter(self.completer)
        
        athlete_layout.addWidget(self.search_input)
        athlete_group.setLayout(athlete_layout)
        layout.addWidget(athlete_group)
        
        # Ödeme detayları grubu
        payment_group = QGroupBox("Ödeme Detayları")
        payment_layout = QFormLayout()
        
        # Tarih seçici
        self.date_input = QDateEdit()
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        self.date_input.setDisplayFormat("dd/MM/yyyy")
        
        # Ay seçici
        self.payment_month_combo = QComboBox()
        self.payment_month_combo.addItems([
            "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"
        ])
        self.payment_month_combo.setCurrentIndex(QDate.currentDate().month() - 1)
        
        # Tutar bilgileri
        self.current_fee_label = QLabel()
        self.amount_input = QLineEdit()
        self.amount_input.setReadOnly(True)
        
        payment_layout.addRow("Ödeme Tarihi:", self.date_input)
        payment_layout.addRow("Ödenen Ay:", self.payment_month_combo)
        payment_layout.addRow("Normal Aidat:", self.current_fee_label)
        payment_layout.addRow("Ödenecek Tutar:", self.amount_input)
        
        payment_group.setLayout(payment_layout)
        layout.addWidget(payment_group)
        
        # Butonlar
        button_layout = QHBoxLayout()
        save_btn = QPushButton("Ödemeyi Kaydet")
        save_btn.setStyleSheet("background-color: #4CAF50; color: white; padding: 5px;")
        cancel_btn = QPushButton("İptal")
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Bağlantılar
        self.search_input.textChanged.connect(self.on_search_changed)
        save_btn.clicked.connect(self.validate_and_accept)
        cancel_btn.clicked.connect(self.reject)
        self.date_input.dateChanged.connect(self.calculate_payment)
        self.completer.activated[str].connect(self.on_athlete_selected)
        
        # Öğrencileri yükle
        self.load_students()

    def load_students(self):
        """Öğrencileri yükle ve completer'a ekle"""
        try:
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT id, name || ' ' || surname as full_name, fee, registration_date 
                    FROM students 
                    ORDER BY name
                """)
                self.all_athletes = []
                
                for student in cursor.fetchall():
                    self.all_athletes.append({
                        'id': student[0],
                        'name': student[1],
                        'fee': student[2],
                        'reg_date': student[3]
                    })
                
                # Completer için sporcu isimlerini ayarla
                athlete_names = [athlete['name'] for athlete in self.all_athletes]
                self.completer.setModel(QStringListModel(athlete_names))
                    
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Öğrenciler yüklenirken hata: {str(e)}")

    def on_search_changed(self, text):
        """Arama metni değiştiğinde"""
        if not text:
            self.current_fee_label.clear()
            self.amount_input.clear()

    def on_athlete_selected(self, name):
        """Sporcu seçildiğinde"""
        self.selected_athlete = next(
            (athlete for athlete in self.all_athletes if athlete['name'] == name),
            None
        )
        if self.selected_athlete:
            self.calculate_payment()
    
    def calculate_payment(self):
        """Ödeme tutarını hesapla"""
        if hasattr(self, 'selected_athlete'):
            try:
                tam_aidat = self.selected_athlete['fee']
                odeme_tarihi = self.date_input.date().toPyDate()
                
                # Normal aidat tutarını göster
                self.current_fee_label.setText(f"{tam_aidat} TL")
                self.amount_input.setText(str(tam_aidat))
                
            except Exception as e:
                QMessageBox.warning(self, "Hata", f"Hesaplama hatası: {str(e)}")

    def validate_and_accept(self):
        try:
            if not hasattr(self, 'selected_athlete'):
                QMessageBox.warning(self, "Hata", "Lütfen bir sporcu seçin!")
                return
                
            if not self.amount_input.text():
                QMessageBox.warning(self, "Hata", "Ödeme tutarı hesaplanamadı!")
                return
                
            athlete_id = self.selected_athlete['id']
            payment_date = self.date_input.date().toPyDate()
            amount = float(self.amount_input.text())
            athlete_name = self.selected_athlete['name']
            payment_month = self.payment_month_combo.currentIndex() + 1
            payment_year = payment_date.year
            
            # Mükerrer ödeme kontrolü
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT COUNT(*) FROM payments 
                    WHERE student_id = ? AND payment_month = ? AND payment_year = ?
                """, (athlete_id, payment_month, payment_year))
                
                if cursor.fetchone()[0] > 0:
                    raise ValueError("Bu ay için zaten ödeme yapılmış!")
                
                # Ödemeyi kaydet
                cursor.execute("""
                    INSERT INTO payments (
                        student_id, amount, payment_date, 
                        payment_month, payment_year, status
                    ) VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    athlete_id,
                    amount,
                    payment_date.strftime('%Y-%m-%d'),
                    payment_month,
                    payment_year,
                    'Ödendi'
                ))
                
                # Sporcunun ödeme durumunu güncelle
                cursor.execute("""
                    UPDATE students 
                    SET payment_status = 'Ödendi'
                    WHERE id = ?
                """, (athlete_id,))
            
            # Tüm tabloları yenile
            main_window = self.parent().parent
            main_window.refresh_all_tabs()
            
            QMessageBox.information(
                self,
                "Başarılı",
                f"Ödeme başarıyla kaydedildi!\n\n"
                f"Sporcu: {athlete_name}\n"
                f"Tutar: {amount} TL\n"
                f"Dönem: {payment_month}/{payment_year}\n"
                f"Tarih: {payment_date.strftime('%d/%m/%Y')}"
            )
            
            self.accept()
            
        except ValueError as e:
            QMessageBox.warning(self, "Hata", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Beklenmeyen bir hata oluştu: {str(e)}")

    def on_student_selected(self, index):
        """Sporcu seçildiğinde"""
        student_id = self.student_combo.currentData()
        if student_id:
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT id, name || ' ' || surname as full_name, fee, registration_date 
                    FROM students 
                    WHERE id = ?
                """, (student_id,))
                student = cursor.fetchone()
                
                if student:
                    self.selected_athlete = {
                        'id': student[0],
                        'name': student[1],
                        'fee': student[2],
                        'reg_date': student[3]
                    }
                    # Seçilen sporcunun aidat tutarını göster
                    self.current_fee_label.setText(f"{student[2]} TL")
                    self.amount_input.setText(str(student[2]))
                    self.calculate_payment()

class EquipmentTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.init_ui()
        self.load_data()  # Başlangıçta verileri yükle
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Üst butonlar
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton("Yeni Malzeme Ekle")
        stock_in_btn = QPushButton("Stok Girişi")
        give_equipment_btn = QPushButton("Malzeme Ver")
        delete_btn = QPushButton("Sil")  # Silme butonu ekle
        
        button_layout.addWidget(add_btn)
        button_layout.addWidget(stock_in_btn)
        button_layout.addWidget(give_equipment_btn)
        button_layout.addWidget(delete_btn)  # Silme butonu eklendi
        button_layout.addStretch()
        
        layout.addLayout(button_layout)
        
        # Malzeme tablosu
        self.equipment_table = QTableWidget()
        self.equipment_table.setColumnCount(8)  # Sütun sayısını 8'e çıkar
        self.equipment_table.setHorizontalHeaderLabels([
            "ID", "Kategori", "Malzeme Adı", "Beden", 
            "Alış Fiyatı", "Satış Fiyatı", "Stok", "Min. Stok"
        ])
        self.equipment_table.verticalHeader().setVisible(False)
        self.equipment_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # Sütun genişliklerini ayarla
        self.equipment_table.setColumnWidth(0, 50)   # ID
        self.equipment_table.setColumnWidth(1, 100)  # Kategori
        self.equipment_table.setColumnWidth(2, 150)  # Malzeme Adı
        self.equipment_table.setColumnWidth(3, 80)   # Beden
        self.equipment_table.setColumnWidth(4, 100)  # Alış Fiyatı
        self.equipment_table.setColumnWidth(5, 100)  # Satış Fiyatı
        self.equipment_table.setColumnWidth(6, 80)   # Stok
        self.equipment_table.setColumnWidth(7, 80)   # Min. Stok
        
        layout.addWidget(self.equipment_table)
        
        self.setLayout(layout)
        
        # Bağlantılar
        add_btn.clicked.connect(self.add_equipment)
        stock_in_btn.clicked.connect(self.add_stock)
        give_equipment_btn.clicked.connect(self.give_equipment)
        delete_btn.clicked.connect(self.delete_equipment)  # Silme fonksiyonu bağlantısı
    
    def load_data(self):
        try:
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT 
                        e.id, c.name as category, e.name, e.size,
                        e.purchase_price, e.sale_price, e.stock_quantity, e.min_stock_level
                    FROM equipment e
                    LEFT JOIN equipment_categories c ON e.category_id = c.id
                    ORDER BY c.name, e.name
                """)
                
                equipment = cursor.fetchall()
                
                self.equipment_table.setRowCount(0)
                for row, data in enumerate(equipment):
                    self.equipment_table.insertRow(row)
                    for col, value in enumerate(data):
                        if col in [4, 5]:  # Fiyat sütunları
                            item = QTableWidgetItem(f"{value:.2f} TL")
                        else:
                            item = QTableWidgetItem(str(value))
                        
                        if col == 6 and value <= data[7]:  # Stok miktarı minimum seviyenin altında
                            item.setBackground(QColor("#FFCDD2"))
                            
                        self.equipment_table.setItem(row, col, item)
                        
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Veriler yüklenirken hata oluştu: {str(e)}")

    def add_equipment(self):
        dialog = EquipmentDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
    
    def add_stock(self):
        dialog = StockDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
    
    def give_equipment(self):
        dialog = GiveEquipmentDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()

    def delete_equipment(self):
        """Seçili malzemeyi sil"""
        try:
            current_row = self.equipment_table.currentRow()
            if current_row < 0:
                QMessageBox.warning(self, "Uyarı", "Lütfen silinecek malzemeyi seçin!")
                return
            
            # Seçili malzemenin bilgilerini al
            equipment_id = self.equipment_table.item(current_row, 0).text()
            equipment_name = self.equipment_table.item(current_row, 2).text()
            
            # Kullanıcıdan onay al
            reply = QMessageBox.question(
                self,
                "Onay",
                f"{equipment_name} malzemesini silmek istediğinizden emin misiniz?\n\n"
                "Bu işlem geri alınamaz!",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                with DatabaseConnection() as cursor:
                    # Önce bağlı kayıtları sil
                    cursor.execute("DELETE FROM stock_movements WHERE equipment_id = ?", (equipment_id,))
                    cursor.execute("DELETE FROM student_equipment WHERE equipment_id = ?", (equipment_id,))
                    
                    # Sonra malzemeyi sil
                    cursor.execute("DELETE FROM equipment WHERE id = ?", (equipment_id,))
                
                # Tabloyu güncelle
                self.load_data()
                QMessageBox.information(self, "Başarılı", "Malzeme başarıyla silindi.")
                
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Malzeme silinirken hata oluştu: {str(e)}")

# Veritabanı bağlantısını optimize et
def get_db_connection():
    conn = sqlite3.connect('korfez_spor.db')
    conn.execute('PRAGMA journal_mode=WAL')  # Write-Ahead Logging
    conn.execute('PRAGMA synchronous=NORMAL')  # Daha hızlı yazma
    conn.execute('PRAGMA cache_size=10000')  # Cache boyutunu artır
    return conn

def import_excel_modules():
    """Excel işlemleri için gerekli modülleri import eder"""
    try:
        global pd, openpyxl
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        return True
    except ImportError as e:
        QMessageBox.critical(
            None,
            "Modül Hatası",
            f"Excel işlemleri için gerekli modüller yüklenemedi!\n\n"
            f"Lütfen şu komutları çalıştırın:\n"
            f"pip install pandas\n"
            f"pip install openpyxl\n\n"
            f"Hata: {str(e)}"
        )
        return False

class EquipmentDialog(QDialog):
    def __init__(self, parent=None, equipment_id=None):
        super().__init__(parent)
        self.equipment_id = equipment_id
        self.setWindowTitle("Malzeme Ekle/Düzenle")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        
        # Malzeme bilgileri
        form_layout = QFormLayout()
        
        # Kategori seçimi
        self.category_combo = QComboBox()
        self.load_categories()
        
        # Diğer alanlar
        self.name_input = QLineEdit()
        self.size_input = QLineEdit()
        self.min_stock_input = QSpinBox()
        self.min_stock_input.setMinimum(1)
        self.min_stock_input.setValue(5)
        
        # Fiyat alanları
        self.purchase_price_input = QLineEdit()
        self.purchase_price_input.setValidator(QDoubleValidator(0.00, 9999.99, 2))
        self.sale_price_input = QLineEdit()
        self.sale_price_input.setValidator(QDoubleValidator(0.00, 9999.99, 2))
        
        form_layout.addRow("Kategori:", self.category_combo)
        form_layout.addRow("Malzeme Adı:", self.name_input)
        form_layout.addRow("Beden:", self.size_input)
        form_layout.addRow("Alış Fiyatı (TL):", self.purchase_price_input)
        form_layout.addRow("Satış Fiyatı (TL):", self.sale_price_input)
        form_layout.addRow("Minimum Stok:", self.min_stock_input)
        
        layout.addLayout(form_layout)
        
        # Butonlar
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
        
        # Eğer düzenleme ise verileri yükle
        if equipment_id:
            self.load_equipment_data()
    
    def load_categories(self):
        """Kategorileri combobox'a yükle"""
        try:
            with DatabaseConnection() as cursor:
                cursor.execute("SELECT id, name FROM equipment_categories ORDER BY name")
                categories = cursor.fetchall()
                
                self.category_combo.clear()
                self.category_combo.addItem("Kategori Seçin", None)
                for category_id, name in categories:
                    self.category_combo.addItem(name, category_id)
                    
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kategoriler yüklenirken hata: {str(e)}")
    
    def load_equipment_data(self):
        """Mevcut malzeme bilgilerini yükle"""
        try:
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT category_id, name, size, purchase_price, sale_price, min_stock_level
                    FROM equipment WHERE id = ?
                """, (self.equipment_id,))
                
                equipment = cursor.fetchone()
                if equipment:
                    self.category_combo.setCurrentIndex(
                        self.category_combo.findData(equipment[0])
                    )
                    self.name_input.setText(equipment[1])
                    self.size_input.setText(equipment[2])
                    self.purchase_price_input.setText(str(equipment[3]))
                    self.sale_price_input.setText(str(equipment[4]))
                    self.min_stock_input.setValue(equipment[5])
                    
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Malzeme bilgileri yüklenirken hata: {str(e)}")
    
    def validate_and_accept(self):
        """Verileri kontrol et ve kaydet"""
        try:
            # Validasyonlar
            if not self.category_combo.currentData():
                raise ValueError("Lütfen bir kategori seçin!")
                
            if not self.name_input.text().strip():
                raise ValueError("Malzeme adı boş olamaz!")
                
            if not self.purchase_price_input.text():
                raise ValueError("Lütfen alış fiyatı girin!")
            
            if not self.sale_price_input.text():
                raise ValueError("Lütfen satış fiyatı girin!")
            
            # Veritabanına kaydet
            with DatabaseConnection() as cursor:
                if self.equipment_id:  # Güncelleme
                    cursor.execute("""
                        UPDATE equipment SET
                            category_id = ?,
                            name = ?,
                            size = ?,
                            purchase_price = ?,
                            sale_price = ?,
                            min_stock_level = ?
                        WHERE id = ?
                    """, (
                        self.category_combo.currentData(),
                        self.name_input.text().strip(),
                        self.size_input.text().strip(),
                        float(self.purchase_price_input.text()),
                        float(self.sale_price_input.text()),
                        self.min_stock_input.value(),
                        self.equipment_id
                    ))
                else:  # Yeni kayıt
                    cursor.execute("""
                        INSERT INTO equipment (
                            category_id, name, size, purchase_price,
                            sale_price, min_stock_level, stock_quantity
                        ) VALUES (?, ?, ?, ?, ?, ?, 0)
                    """, (
                        self.category_combo.currentData(),
                        self.name_input.text().strip(),
                        self.size_input.text().strip(),
                        float(self.purchase_price_input.text()),
                        float(self.sale_price_input.text()),
                        self.min_stock_input.value()
                    ))
            
            self.accept()
            
        except ValueError as e:
            QMessageBox.warning(self, "Hata", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Kayıt sırasında hata: {str(e)}")

class StockDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Stok Giriş/Çıkış")
        self.setModal(True)
        self.setMinimumWidth(400)
        
        layout = QVBoxLayout()
        
        # Malzeme seçimi
        form_layout = QFormLayout()
        
        # Malzeme seçim combobox'ı
        self.equipment_combo = QComboBox()
        self.load_equipment()
        
        # İşlem tipi seçimi
        self.movement_type = QComboBox()
        self.movement_type.addItems(["Giriş", "Çıkış"])
        
        # Miktar girişi
        self.quantity_input = QSpinBox()
        self.quantity_input.setMinimum(1)
        self.quantity_input.setMaximum(9999)
        
        # Açıklama
        self.description_input = QTextEdit()
        self.description_input.setMaximumHeight(60)
        self.description_input.setPlaceholderText("İşlem açıklaması...")
        
        form_layout.addRow("Malzeme:", self.equipment_combo)
        form_layout.addRow("İşlem Tipi:", self.movement_type)
        form_layout.addRow("Miktar:", self.quantity_input)
        form_layout.addRow("Açıklama:", self.description_input)
        
        layout.addLayout(form_layout)
        
        # Mevcut stok bilgisi
        self.current_stock_label = QLabel()
        layout.addWidget(self.current_stock_label)
        
        # Butonlar
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
        
        # Bağlantılar
        self.equipment_combo.currentIndexChanged.connect(self.update_stock_info)
        
    def load_equipment(self):
        """Malzemeleri combobox'a yükle"""
        try:
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT e.id, e.name, c.name, e.size, e.stock_quantity
                    FROM equipment e
                    LEFT JOIN equipment_categories c ON e.category_id = c.id
                    ORDER BY c.name, e.name
                """)
                equipment = cursor.fetchall()
                
                self.equipment_combo.clear()
                self.equipment_combo.addItem("Malzeme Seçin", None)
                
                for eq in equipment:
                    self.equipment_combo.addItem(
                        f"{eq[2]} - {eq[1]} ({eq[3]}) - Stok: {eq[4]}", 
                        {"id": eq[0], "stock": eq[4]}
                    )
                    
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Malzemeler yüklenirken hata: {str(e)}")
    
    def update_stock_info(self):
        """Seçilen malzemenin stok bilgisini göster"""
        data = self.equipment_combo.currentData()
        if data:
            self.current_stock_label.setText(f"Mevcut Stok: {data['stock']}")
        else:
            self.current_stock_label.clear()
    
    def validate_and_accept(self):
        """Verileri kontrol et ve kaydet"""
        try:
            equipment_data = self.equipment_combo.currentData()
            if not equipment_data:
                raise ValueError("Lütfen bir malzeme seçin!")
            
            quantity = self.quantity_input.value()
            movement_type = "IN" if self.movement_type.currentText() == "Giriş" else "OUT"
            
            # Çıkış işleminde stok kontrolü
            if movement_type == "OUT" and quantity > equipment_data['stock']:
                raise ValueError("Stokta yeterli malzeme yok!")
            
            with DatabaseConnection() as cursor:
                # Stok hareketini kaydet
                cursor.execute("""
                    INSERT INTO stock_movements (
                        equipment_id, movement_type, quantity,
                        date, description
                    ) VALUES (?, ?, ?, ?, ?)
                """, (
                    equipment_data['id'],
                    movement_type,
                    quantity,
                    datetime.now().strftime('%Y-%m-%d'),
                    self.description_input.toPlainText().strip()
                ))
                
                # Stok miktarını güncelle
                new_quantity = equipment_data['stock'] + quantity if movement_type == "IN" else equipment_data['stock'] - quantity
                
                cursor.execute("""
                    UPDATE equipment 
                    SET stock_quantity = ?
                    WHERE id = ?
                """, (new_quantity, equipment_data['id']))

                # Stok girişi ise gider olarak kaydet
                if movement_type == "IN":
                    # Malzeme alış fiyatını al
                    cursor.execute("SELECT purchase_price FROM equipment WHERE id = ?", (equipment_data['id'],))
                    unit_price = cursor.fetchone()[0]
                    total_cost = unit_price * quantity

                    # Gider olarak kaydet
                    cursor.execute("""
                        INSERT INTO expenses (
                            date, category, description, amount
                        ) VALUES (?, ?, ?, ?)
                    """, (
                        datetime.now().strftime('%Y-%m-%d'),
                        "Malzeme Alımı",
                        f"{self.equipment_combo.currentText()} - {quantity} adet",
                        total_cost
                    ))

                    # Muhasebe tablosunu güncelle
                    main_window = self.parent().parent
                    if hasattr(main_window, 'accounting_tab'):
                        main_window.accounting_tab.load_data()

            self.accept()
            
        except ValueError as e:
            QMessageBox.warning(self, "Hata", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"İşlem sırasında hata: {str(e)}")

class GiveEquipmentDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Sporcu Malzeme Ver")
        self.setModal(True)
        self.setMinimumWidth(500)
        
        layout = QVBoxLayout()
        
        # Sporcu seçimi
        student_group = QGroupBox("Sporcu Seçimi")
        student_layout = QVBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Sporcu adını yazın...")
        
        # Completer için model
        self.completer = QCompleter()
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.completer.setFilterMode(Qt.MatchContains)
        self.search_input.setCompleter(self.completer)
        
        student_layout.addWidget(self.search_input)
        student_group.setLayout(student_layout)
        layout.addWidget(student_group)
        
        # Malzeme seçimi
        equipment_group = QGroupBox("Malzeme Seçimi")
        equipment_layout = QFormLayout()
        
        self.equipment_combo = QComboBox()
        self.load_equipment()
        
        self.payment_status = QComboBox()
        self.payment_status.addItems(["Ödendi", "Ödenmedi"])
        
        equipment_layout.addRow("Malzeme:", self.equipment_combo)
        equipment_layout.addRow("Ödeme Durumu:", self.payment_status)
        
        # Fiyat bilgisi
        self.price_label = QLabel()
        equipment_layout.addRow("Fiyat:", self.price_label)
        
        equipment_group.setLayout(equipment_layout)
        layout.addWidget(equipment_group)
        
        # Butonlar
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
        
        # Bağlantılar
        self.load_students()
        self.equipment_combo.currentIndexChanged.connect(self.update_price)
        self.completer.activated[str].connect(self.on_student_selected)
        
    def load_students(self):
        """Öğrencileri yükle"""
        try:
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT id, name || ' ' || surname as full_name
                    FROM students 
                    ORDER BY name
                """)
                self.all_students = cursor.fetchall()
                
                # Completer için isimleri ayarla
                student_names = [student[1] for student in self.all_students]
                self.completer.setModel(QStringListModel(student_names))
                
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Öğrenciler yüklenirken hata: {str(e)}")
    
    def load_equipment(self):
        """Malzemeleri combobox'a yükle"""
        try:
            with DatabaseConnection() as cursor:
                cursor.execute("""
                    SELECT e.id, e.name, c.name, e.size, e.stock_quantity, e.sale_price
                    FROM equipment e
                    LEFT JOIN equipment_categories c ON e.category_id = c.id
                    WHERE e.stock_quantity > 0
                    ORDER BY c.name, e.name
                """)
                equipment = cursor.fetchall()
                
                self.equipment_combo.clear()
                self.equipment_combo.addItem("Malzeme Seçin", None)
                
                for eq in equipment:
                    self.equipment_combo.addItem(
                        f"{eq[2]} - {eq[1]} ({eq[3]}) - Stok: {eq[4]}", 
                        {"id": eq[0], "sale_price": eq[5]}  # price yerine sale_price kullan
                    )
                    
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Malzemeler yüklenirken hata: {str(e)}")
    
    def update_price(self):
        """Seçilen malzemenin fiyatını göster"""
        data = self.equipment_combo.currentData()
        if data:
            self.price_label.setText(f"{data['sale_price']:.2f} TL")  # price yerine sale_price kullan
        else:
            self.price_label.clear()
    
    def on_student_selected(self, name):
        """Sporcu seçildiğinde"""
        self.selected_student = next(
            (student for student in self.all_students if student[1] == name),
            None
        )
    
    def validate_and_accept(self):
        """Verileri kontrol et ve kaydet"""
        try:
            if not hasattr(self, 'selected_student'):
                raise ValueError("Lütfen bir sporcu seçin!")
            
            equipment_data = self.equipment_combo.currentData()
            if not equipment_data:
                raise ValueError("Lütfen bir malzeme seçin!")
            
            with DatabaseConnection() as cursor:
                # Malzeme kaydını ekle
                cursor.execute("""
                    INSERT INTO student_equipment (
                        student_id, equipment_id, given_date,
                        payment_status, payment_amount
                    ) VALUES (?, ?, ?, ?, ?)
                """, (
                    self.selected_student[0],  # student_id
                    equipment_data['id'],      # equipment_id
                    datetime.now().strftime('%Y-%m-%d'),
                    self.payment_status.currentText(),
                    equipment_data['sale_price']
                ))
                
                # Stok miktarını güncelle
                cursor.execute("""
                    UPDATE equipment 
                    SET stock_quantity = stock_quantity - 1
                    WHERE id = ?
                """, (equipment_data['id'],))
                
                # Ödendi ise gelir olarak ekle
                if self.payment_status.currentText() == "Ödendi":
                    cursor.execute("""
                        INSERT INTO income (
                            date, type, description, amount
                        ) VALUES (?, 'Malzeme', ?, ?)
                    """, (
                        datetime.now().strftime('%Y-%m-%d'),
                        f"Malzeme Satışı - {self.equipment_combo.currentText()}",
                        equipment_data['sale_price']
                    ))
            
            self.accept()
            
        except ValueError as e:
            QMessageBox.warning(self, "Hata", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"İşlem sırasında hata: {str(e)}")

def validate_tc_no(tc_no):
    return len(tc_no) == 11 and tc_no.isdigit()

def validate_phone(phone):
    return phone.isdigit() and not phone.startswith('0')

def safe_execute(cursor, query, params):
    try:
        return cursor.execute(query, params)
    except sqlite3.Error as e:
        log_error(f"Database error: {e}")
        raise

def calculate_fee(base_fee, payment_date, registration_date):
    if payment_date.month == registration_date.month:
        days_in_month = calendar.monthrange(payment_date.year, payment_date.month)[1]
        remaining_days = days_in_month - registration_date.day + 1
        return round((base_fee / days_in_month) * remaining_days, 2)
    return base_fee

def check_stock_level(equipment_id, quantity):
    with DatabaseConnection() as cursor:
        cursor.execute("SELECT stock_quantity FROM equipment WHERE id = ?", (equipment_id,))
        current_stock = cursor.fetchone()[0]
        if current_stock < quantity:
            raise ValueError(f"Yetersiz stok! Mevcut: {current_stock}, İstenen: {quantity}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # Uygulama stili
    app.setStyle('Fusion')
    
    # Ana pencereyi oluştur
    window = MainWindow()
    window.show()
    
    # Uygulamayı başlat
    sys.exit(app.exec_()) 